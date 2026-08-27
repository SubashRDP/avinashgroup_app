# Copyright (c) 2026, Raindrop and contributors
# For license information, please see license.txt

import json
import os

import frappe
from frappe import _


def _as_list(value):
	"""Normalize a MultiSelectList/Link filter value (list, JSON string, or single) to a list."""
	if not value:
		return []
	if isinstance(value, str):
		value = value.strip()
		if value.startswith("["):
			try:
				value = json.loads(value)
			except Exception:
				return [value]
		else:
			return [value]
	if isinstance(value, (list, tuple, set)):
		return [v for v in value if v]
	return [value]


@frappe.whitelist()
def get_company_branches(company=None, txt=None):
	"""Branch options scoped to the selected company via Branch.custom_company."""
	company = _as_list(company)
	like = f"%{(txt or '').strip()}%"
	conditions = ["(b.name LIKE %(txt)s OR b.branch LIKE %(txt)s)"]
	values = {"txt": like}
	if company:
		conditions.append("(b.custom_company IN %(company)s OR COALESCE(b.custom_company, '') = '')")
		values["company"] = tuple(company)
	where = " AND ".join(conditions)

	return frappe.db.sql(
		f"""
		SELECT b.name AS value, b.branch AS label, b.name AS description
		FROM `tabBranch` b
		WHERE {where}
		ORDER BY b.branch
		LIMIT 50
		""",
		values,
		as_dict=True,
	)


@frappe.whitelist()
def get_company_customers(company=None, txt=None):
	"""Customer options scoped to the selected company via the customer's custom_company."""
	company = _as_list(company)
	like = f"%{(txt or '').strip()}%"
	conditions = ["(cust.name LIKE %(txt)s OR cust.customer_name LIKE %(txt)s)"]
	values = {"txt": like}
	if company:
		conditions.append("(cust.custom_company IN %(company)s OR COALESCE(cust.custom_company, '') = '')")
		values["company"] = tuple(company)
	where = " AND ".join(conditions)

	return frappe.db.sql(
		f"""
		SELECT cust.name AS value, cust.customer_name AS label, cust.name AS description
		FROM `tabCustomer` cust
		WHERE {where}
		ORDER BY cust.customer_name
		LIMIT 50
		""",
		values,
		as_dict=True,
	)


def _fmt_inr(v):
	if not v:
		return ''
	import locale
	try:
		locale.setlocale(locale.LC_ALL, 'en_IN.UTF-8')
		return locale.format_string('%.2f', v, grouping=True)
	except Exception:
		return '{:,.2f}'.format(v)


def _get_customer_display(filters):
	if filters.get('customer'):
		rows = frappe.db.get_all('Customer',
			filters=[['name', 'in', filters.get('customer')]],
			fields=['customer_name'], order_by='customer_name asc')
		return ', '.join(r.customer_name for r in rows)
	return ''


def _get_govt_heading_info(filters):
	"""PAN, company name(s), and BS fiscal year ("83/84") for the govt heading block —
	resolved here (not in the template/JS) so the print and on-screen views agree."""
	companies = filters.get('company') or []
	if isinstance(companies, str):
		companies = [companies]

	pan = ''
	company_name = ''
	fiscal_year = ''
	if companies:
		rows = frappe.db.get_all('Company',
			filters=[['name', 'in', companies]],
			fields=['name', 'tax_id'], order_by='name asc')
		pan = ', '.join(r.tax_id for r in rows if r.tax_id)
		company_name = ', '.join(r.name for r in rows)

		if filters.get('to_date'):
			from avinashgroup_app.custom_code.CBMS.utils import cbms_fiscal_year
			try:
				fiscal_year = cbms_fiscal_year(filters.get('to_date'), company=companies[0])
			except Exception:
				fiscal_year = ''

	return {'pan': pan, 'company_name': company_name, 'fiscal_year': fiscal_year}


# Nepali-only labelling for the two official govt VAT books (Rule 23(1)(ja)): "बिक्री खाता"
# (Sales Book, Is Return unticked) and "बिक्री फिर्ता खाता" (Sales Return Book, Is Return
# ticked). Only fields with a govt Nepali equivalent are shown; Date and Branch have none
# and are dropped entirely (not merely relabelled) — dropping them makes miti/bill_no/
# customer/vat_number contiguous, so बीजक is a genuine 4-column merge, no reordering needed.
# export_declaration_no/miti are placeholder columns the govt form has that this report has
# no data for — always blank, never pulled from Sales Invoice.
def _govt_groups(is_return):
	if is_return:
		return {
			'bijak': 'बीजक',
			'taxable': 'करयोग्य फिर्ता',
		}
	return {
		'bijak': 'बीजक',
		'taxable': 'करयोग्य बिक्री',
		'export': 'निकासी',
	}


def _govt_fields(is_return):
	"""(fieldname, group or None, Nepali label, css align, fieldtype, width)."""
	if is_return:
		return [
			('miti',                  'bijak',   'मिति',                                      'Data',     120),
			('bill_no',               'bijak',   'बीजक नम्बर',                                'Data',     170),
			('customer',              'bijak',   'खरिदकर्ताको नाम',                            'Data',     180),
			('vat_number',            'bijak',   'खरिदकर्ताको स्थायी लेखा नम्बर',              'Data',     150),
			('item_description',      None,      'वस्तु वा सेवाको नाम',                        'Data',     220),
			('qty',                   None,      'वस्तु वा सेवाको परिमाण',                     'Float',    110),
			('uom',                   None,      'वस्तु वा सेवाको एकाई',                       'Data',     100),
			('total_sales',           None,      'जम्मा फिर्ता (रु)',                           'Currency', 140),
			('tax_free_sale',         None,      'स्थानीय कर छुटको फिर्ता मूल्य (रु)',          'Currency', 160),
			('taxable_sales',         'taxable', 'मूल्य (रु)',                                 'Currency', 120),
			('vat',                   'taxable', 'कर (रु)',                                    'Currency', 110),
		]
	return [
		('miti',                     'bijak',  'मिति',                                        'Data',     120),
		('bill_no',                  'bijak',  'बीजक नम्बर',                                  'Data',     170),
		('customer',                 'bijak',  'खरिदकर्ताको नाम',                              'Data',     180),
		('vat_number',               'bijak',  'खरिदकर्ताको स्थायी लेखा नम्बर',                'Data',     150),
		('total_sales',              None,     'जम्मा बिक्री / निकासी (रु)',                   'Currency', 140),
		('tax_free_sale',            None,     'स्थानीय कर छुटको बिक्री मूल्य (रु)',           'Currency', 160),
		('taxable_sales',            'taxable', 'मूल्य (रु)',                                  'Currency', 120),
		('vat',                      'taxable', 'कर (रु)',                                     'Currency', 110),
		('export_npr',               'export',  'निकासी गरेको वस्तु वा सेवाको मूल्य (रु)',      'Currency', 160),
		('export_country',           'export',  'निकासी गरेको देश',                            'Data',     130),
		('export_declaration_no',    'export',  'निकासी प्रज्ञापनपत्र नम्बर',                   'Data',     140),
		('export_declaration_miti',  'export',  'निकासी प्रज्ञापनपत्र मिति',                    'Data',     130),
	]


def get_govt_columns(is_return=False):
	"""Column defs (Nepali labels) for the on-screen/print govt view — Sales Book (default)
	or Sales Return Book (is_return)."""
	return [
		{"fieldname": f, "label": label, "fieldtype": ftype, "width": width}
		for f, _group, label, ftype, width in _govt_fields(is_return)
	]


def _build_govt_layout(selected_columns, is_return=False):
	"""Row1 (group super-headers only) for the govt book layout. Individual (non-grouped)
	sub labels come from get_govt_columns() as the real column headers, not from this overlay."""
	show_all = not selected_columns
	visible = lambda f: show_all or f in selected_columns
	groups = _govt_groups(is_return)

	body_fields = [
		{'key': f, 'group': group, 'sub': label, 'css': 'l' if ftype == 'Data' else 'r', 'kind': 'currency' if ftype == 'Currency' else ('bill' if f == 'bill_no' else 'text')}
		for f, group, label, ftype, width in _govt_fields(is_return)
		if visible(f)
	]

	row1 = []
	i = 0
	while i < len(body_fields):
		bf = body_fields[i]
		if bf['group']:
			j = i
			while j < len(body_fields) and body_fields[j]['group'] == bf['group']:
				j += 1
			row1.append({'text': groups[bf['group']], 'css': bf['css'], 'colspan': j - i, 'rowspan': 1})
			i = j
		else:
			row1.append({'text': bf['sub'], 'css': bf['css'], 'colspan': 1, 'rowspan': 2})
			i += 1

	return {'row1': row1, 'body_fields': body_fields}


@frappe.whitelist()
def get_govt_header_html(selected_columns=None, is_return=None):
	"""Row1 (Nepali group super-headers only) as an HTML fragment, for the on-screen report
	to show above its own (already Nepali-labelled) column headers. Standalone (non-merged)
	columns render blank here — their label is the real column header underneath (row 2)."""
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	layout = _build_govt_layout(selected_columns, is_return=bool(frappe.utils.cint(is_return)))

	tds = []
	for c in layout['row1']:
		is_group_cell = c.get('colspan', 1) > 1
		attrs = f' colspan="{c["colspan"]}"' if is_group_cell else ''
		text = c['text'] if is_group_cell else ''
		tds.append(f'<th class="{c["css"]}"{attrs}>{frappe.utils.escape_html(text)}</th>')

	return '<table><tr>' + ''.join(tds) + '</tr></table>'


@frappe.whitelist()
def get_print_html(filters, selected_columns=None, orientation=None):
	if isinstance(filters, str):
		filters = frappe._dict(json.loads(filters))
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	is_return = bool(filters.get('is_return'))
	_, data = execute(filters)
	customer_display = _get_customer_display(filters)
	govt_layout = _build_govt_layout(selected_columns, is_return=is_return)
	govt_heading = _get_govt_heading_info(filters)

	template_path = os.path.join(os.path.dirname(__file__), 'sales_register_report_pdf.html')
	with open(template_path) as f:
		template_content = f.read()

	return frappe.render_template(template_content, {
		'filters': filters,
		'data': data,
		'fmt': _fmt_inr,
		'sc': selected_columns or [],
		'is_html_view': True,
		'orientation': orientation or 'Landscape',
		'customer_display': customer_display,
		'govt_layout': govt_layout,
		'govt_heading': govt_heading,
	})


@frappe.whitelist()
def download_pdf(filters, orientation=None, selected_columns=None, view=None):
	from frappe.utils.pdf import get_pdf

	if isinstance(filters, str):
		filters = frappe._dict(json.loads(filters))
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	orientation = orientation if orientation in ('Portrait', 'Landscape') else 'Landscape'

	is_return = bool(filters.get('is_return'))
	_, data = execute(filters)
	customer_display = _get_customer_display(filters)
	govt_layout = _build_govt_layout(selected_columns, is_return=is_return)
	govt_heading = _get_govt_heading_info(filters)

	template_path = os.path.join(os.path.dirname(__file__), 'sales_register_report_pdf.html')
	with open(template_path) as f:
		template_content = f.read()

	# sc = the columns currently shown in the report (picked in the print dialog, or
	# whatever is left after the user removed columns in the datatable). Empty => show all.
	# Page numbers are rendered inside the template body (manual pagination), so they
	# work on plain/unpatched wkhtmltopdf. No --footer-html here: that feature needs the
	# patched-Qt build and is silently ignored otherwise.
	html = frappe.render_template(template_content, {
		'filters': filters,
		'data': data,
		'fmt': _fmt_inr,
		'sc': selected_columns or [],
		'orientation': orientation,
		'customer_display': customer_display,
		'govt_layout': govt_layout,
		'govt_heading': govt_heading,
	})

	options = {
		'page-size': 'A4',
		'orientation': orientation,
		'margin-top': '10mm',
		'margin-right': '10mm',
		'margin-bottom': '15mm',
		'margin-left': '10mm',
		'encoding': 'UTF-8',
		'enable-local-file-access': None,
	}
	pdf_data = get_pdf(html, options)

	frappe.response.filename = 'sales_register.pdf'
	frappe.response.filecontent = pdf_data
	# view=1 (Print) → open inline in the browser tab; otherwise download the file.
	frappe.response.type = 'pdf' if frappe.utils.cint(view) else 'download'


@frappe.whitelist()
def download_excel(filters, selected_columns=None):
	"""XLSX with the same govt VAT book heading/column-group layout as the PDF (title,
	PAN/नाम/साल line, merged group headers) — the stock Export menu can't produce this
	(flat columns + optional filter list only), so this routes through it via export_report()."""
	from io import BytesIO

	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, Side, Border
	from openpyxl.utils import get_column_letter
	from frappe.desk.utils import provide_binary_file

	if isinstance(filters, str):
		filters = frappe._dict(json.loads(filters))
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	_, data = execute(filters)
	is_return = bool(filters.get('is_return'))

	wb = Workbook()
	ws = wb.active
	ws.title = 'Sales Register'

	thin = Side(style='thin', color='999999')
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	center = Alignment(horizontal='center', vertical='center', wrap_text=True)
	left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
	right_align = Alignment(horizontal='right', vertical='center')

	layout = _build_govt_layout(selected_columns, is_return=is_return)
	groups = _govt_groups(is_return)
	heading = _get_govt_heading_info(filters)
	body_fields = layout['body_fields']
	total_cols = len(body_fields)

	title = 'बिक्री फिर्ता खाता' if is_return else 'बिक्री खाता'
	c = ws.cell(row=1, column=1, value=title)
	c.font = Font(bold=True, size=16)
	c.alignment = center
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

	c = ws.cell(row=2, column=1, value='(नियम २३ को उपनियम (१) को खण्ड  (ज) संग सम्बन्धित )')
	c.alignment = center
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

	pan_line = (
		f"करदाता दर्ता नं (PAN) : {heading['pan']}      "
		f"करदाताको नाम: {heading['company_name']}      "
		f"साल  {heading['fiscal_year']}      "
		f"कर अवधि: "
	)
	c = ws.cell(row=4, column=1, value=pan_line)
	c.font = Font(bold=True)
	c.alignment = left_align
	ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)

	group_row, sub_row = 5, 6
	i, col_idx = 0, 1
	while i < len(body_fields):
		bf = body_fields[i]
		if bf['group']:
			j = i
			while j < len(body_fields) and body_fields[j]['group'] == bf['group']:
				j += 1
			span = j - i
			c = ws.cell(row=group_row, column=col_idx, value=groups[bf['group']])
			c.font, c.alignment, c.border = Font(bold=True), center, border
			ws.merge_cells(start_row=group_row, start_column=col_idx, end_row=group_row, end_column=col_idx + span - 1)
			for k in range(span):
				sc = ws.cell(row=sub_row, column=col_idx + k, value=body_fields[i + k]['sub'])
				sc.font, sc.alignment, sc.border = Font(bold=True), center, border
			i, col_idx = j, col_idx + span
		else:
			c = ws.cell(row=group_row, column=col_idx, value=bf['sub'])
			c.font, c.alignment, c.border = Font(bold=True), center, border
			ws.merge_cells(start_row=group_row, start_column=col_idx, end_row=sub_row, end_column=col_idx)
			i, col_idx = i + 1, col_idx + 1

	row = sub_row + 1
	for d in data:
		for ci, bf in enumerate(body_fields, start=1):
			val = d.get(bf['key'])
			c = ws.cell(row=row, column=ci, value=val)
			c.border = border
			if bf['kind'] == 'currency':
				c.alignment = right_align
				c.number_format = '#,##0.00'
			else:
				c.alignment = left_align
			if d.get('bold'):
				c.font = Font(bold=True)
		row += 1

	for i in range(1, total_cols + 1):
		ws.column_dimensions[get_column_letter(i)].width = 18

	buf = BytesIO()
	wb.save(buf)
	provide_binary_file('sales_register', 'xlsx', buf.getvalue())


def execute(filters=None):
	filters = filters or {}
	is_return = bool(filters.get("is_return"))
	# Always the official govt VAT book view now — "बिक्री खाता" (Sales) when Is Return is
	# unticked, "बिक्री फिर्ता खाता" (Sales Return) when ticked. Only fields with a govt
	# Nepali equivalent are shown; there is no more plain flat register.
	columns = get_govt_columns(is_return=is_return)

	# Company is a required filter (see the report JS), so the UI blocks the report
	# with a "Please set the company first" prompt before execute runs. This guard
	# still protects any direct/API call from running the full query with no Company.
	if not _as_list(filters.get("company")):
		return columns, []

	data = get_data(filters)

	# Sales Returns store amounts as negatives in the DB. Show them as positives in the register.
	if is_return and data:
		numeric_fields = [c["fieldname"] for c in columns if c.get("fieldtype") in ("Currency", "Float")]
		for row in data:
			for f in numeric_fields:
				if row.get(f) is not None:
					row[f] = abs(row[f])

	if data:
		total = {"bill_no": _("Total"), "bold": 1}
		for col in columns:
			if col.get("fieldtype") == "Currency":
				total[col["fieldname"]] = sum(row.get(col["fieldname"]) or 0 for row in data)
		data.append(total)
	return columns, data


def get_data(filters):
	is_return = 1 if filters.get("is_return") else 0
	conditions = f"si.docstatus = 1 AND si.is_return = {is_return}"

	if filters.get("company"):
		placeholders = ", ".join(["'{}'".format(c) for c in filters.get("company")])
		conditions += " AND si.company IN ({})".format(placeholders)
	if filters.get("from_date"):
		conditions += " AND si.posting_date >= %(from_date)s"
	if filters.get("to_date"):
		conditions += " AND si.posting_date <= %(to_date)s"
	if filters.get("customer"):
		placeholders = ", ".join(["'{}'".format(c) for c in filters.get("customer")])
		conditions += " AND si.customer IN ({})".format(placeholders)
	if filters.get("branch"):
		placeholders = ", ".join(["'{}'".format(b) for b in filters.get("branch")])
		conditions += " AND si.custom_branch IN ({})".format(placeholders)

	return frappe.db.sql(
		f"""
		SELECT
			si.posting_date                                                                                                               AS date,
			SUBSTRING_INDEX(si.custom_invoice_miti, ' ', 1)                                                                               AS miti,
			COALESCE(si.custom_branch_name, si.name)                                                                                      AS bill_no,
			si.customer_name                                                                                                               AS customer,
			br.branch                                                                                                                      AS branch,
			c.tax_id                                                                                                                       AS vat_number,
			si.custom_total_amount_including_excise                                                                                        AS total_sales,
			SUM(CASE WHEN sii.custom_vat_apply_on = 'VAT 0%%'                                             AND c.territory = 'Nepal'    THEN sii.amount ELSE 0 END) AS tax_free_sale,
			SUM(CASE WHEN sii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND c.territory != 'Nepal' THEN sii.base_amount ELSE 0 END)      AS export_npr,
			SUM(CASE WHEN sii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND c.territory = 'Nepal'  THEN sii.amount ELSE 0 END)           AS taxable_sales,
			SUM(CASE WHEN sii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND c.territory = 'Nepal'  THEN sii.custom_vat_amount ELSE 0 END) AS vat,
			SUM(sii.qty)                                                                                                                     AS qty,
			GROUP_CONCAT(DISTINCT sii.item_name SEPARATOR ', ')                                                                              AS item_description,
			GROUP_CONCAT(DISTINCT sii.uom SEPARATOR ', ')                                                                                    AS uom,
			CASE WHEN c.territory != 'Nepal' THEN c.territory ELSE NULL END                                                                  AS export_country
		FROM `tabSales Invoice` si
		JOIN `tabSales Invoice Item` sii ON sii.parent = si.name
		JOIN `tabCustomer` c ON c.name = si.customer
		LEFT JOIN `tabBranch` br ON br.name = si.custom_branch
		WHERE {conditions}
		GROUP BY si.name
		ORDER BY si.posting_date ASC
		""",
		filters,
		as_dict=True,
	)
