# Copyright (c) 2026, Raindrop and contributors
# For license information, please see license.txt

import json
import os

import frappe
from frappe import _


def _as_list(value):
	"""Normalize a MultiSelectList/Link filter value (list, JSON string, or single) to a list."""
	if not value:
		return []
	if isinstance(value, str):
		value = value.strip()
		if value.startswith("["):
			try:
				value = json.loads(value)
			except Exception:
				return [value]
		else:
			return [value]
	if isinstance(value, (list, tuple, set)):
		return [v for v in value if v]
	return [value]


@frappe.whitelist()
def get_company_suppliers(company=None, txt=None):
	"""Supplier options scoped to the selected company via the supplier's custom_company."""
	company = _as_list(company)
	like = f"%{(txt or '').strip()}%"
	conditions = ["(sup.name LIKE %(txt)s OR sup.supplier_name LIKE %(txt)s)"]
	values = {"txt": like}
	if company:
		conditions.append("(sup.custom_company IN %(company)s OR COALESCE(sup.custom_company, '') = '')")
		values["company"] = tuple(company)
	where = " AND ".join(conditions)

	return frappe.db.sql(
		f"""
		SELECT sup.name AS value, sup.supplier_name AS label, sup.name AS description
		FROM `tabSupplier` sup
		WHERE {where}
		ORDER BY sup.supplier_name
		LIMIT 50
		""",
		values,
		as_dict=True,
	)


def _fmt_inr(v):
	if not v:
		return ''
	import locale
	try:
		locale.setlocale(locale.LC_ALL, 'en_IN.UTF-8')
		return locale.format_string('%.2f', v, grouping=True)
	except Exception:
		return '{:,.2f}'.format(v)


def _get_display_names(filters):
	supplier_display = ''
	ptype_display = ''
	if filters.get('supplier'):
		rows = frappe.db.get_all('Supplier',
			filters=[['name', 'in', filters.get('supplier')]],
			fields=['supplier_name'], order_by='supplier_name asc')
		supplier_display = ', '.join(r.supplier_name for r in rows)
	if filters.get('purchase_type'):
		ptype_display = ', '.join(filters.get('purchase_type'))
	return supplier_display, ptype_display


def _get_govt_heading_info(filters):
	"""PAN, company name(s), the plain BS year ("2083", not the "83/84" fiscal-year label),
	and BS tax period (month name only, e.g. "Shrawan" — the year already shows via साल)
	for the govt heading block — resolved here (not in the template/JS) so the print and
	on-screen views agree."""
	companies = filters.get('company') or []
	if isinstance(companies, str):
		companies = [companies]

	pan = ''
	company_name = ''
	bs_year = ''
	tax_period = ''
	if companies:
		rows = frappe.db.get_all('Company',
			filters=[['name', 'in', companies]],
			fields=['name', 'tax_id'], order_by='name asc')
		pan = ', '.join(r.tax_id for r in rows if r.tax_id)
		company_name = ', '.join(r.name for r in rows)

		if filters.get('to_date'):
			from rdp_common_app.utils.bs_boundaries import ad_to_bs, get_bs_month_name
			try:
				bs = ad_to_bs(frappe.utils.getdate(filters.get('to_date')))
				bs_year = str(bs.year)
				tax_period = get_bs_month_name(bs.month)
			except Exception:
				bs_year = ''
				tax_period = ''

	return {'pan': pan, 'company_name': company_name, 'bs_year': bs_year, 'tax_period': tax_period}


# Nepali-only labelling for the two official govt VAT books (Rule 23(1)(chha)): "खरिद खाता"
# (Purchase Book, Is Return unticked) and "खरिद फिर्ता खाता" (Purchase Return Book, Is Return
# ticked). The report ALWAYS shows one of these two govt layouts now — there is no more
# plain flat register. Only fields with a govt Nepali equivalent are shown; Date, Purchase
# Type, Supplier Invoice No/Date/Miti, Total VAT are dropped, not merely relabelled.
# Dropping them makes voucher_no/pragyanpatra_no/supplier_name/vat_number contiguous again,
# so बीजक/प्रज्ञापनपत्र नम्बर is a genuine 4-column merge — no column reordering needed.
# pragyanpatra_no is a placeholder column the govt form has (Pragyanpatra/declaration no.)
# that this report has no data for — always blank, never pulled from Purchase Invoice.
def _govt_groups(is_return):
	"""Group id -> Nepali super-header spanning its sub-columns."""
	if is_return:
		return {
			'bijak': 'बीजक / प्रज्ञापनपत्र नम्बर',
			'txp': 'करयोग्य फिर्ता (पूंजीगत बाहेक)',
			'tim': 'करयोग्य पैठारी फिर्ता (पूंजीगत बाहेक)',
			'cap': 'पूंजीगत करयोग्य फिर्ता',
		}
	return {
		'bijak': 'बीजक / प्रज्ञापनपत्र नम्बर',
		'txp': 'करयोग्य खरिद (पूंजीगत बाहेक)',
		'tim': 'करयोग्य पैठारी (पूंजीगत बाहेक)',
		'cap': 'पूंजीगत करयोग्य खरिद/पैठारी',
	}


def _govt_fields(is_return):
	"""(fieldname, group or None, Nepali label, css align, fieldtype, width) — the Return
	book adds item description/QTY/UOM (what was returned) and renames the value/tax-free
	labels to "फिर्ता" (return) wording; everything else is identical to the Purchase book."""
	if is_return:
		return [
			('miti',                  None,    'मिति',                                                          'Data',     120),
			('voucher_no',            'bijak', 'बीजक नं.',                                                       'Data',     170),
			('pragyanpatra_no',       'bijak', 'प्रज्ञापनपत्र नं.',                                              'Data',     130),
			('supplier_name',         'bijak', 'आपूर्तिकर्ताको नाम',                                             'Data',     180),
			('vat_number',            'bijak', 'आपूर्तिकर्ताको स्थायी लेखा नम्बर',                                'Data',     150),
			('item_description',      None,    'खरिद/पैठारी फिर्ता गरिएका वस्तु वा सेवाको विवरण',                'Data',     220),
			('qty',                   None,    'खरिद/पैठारी फिर्ता गरिएका वस्तु वा सेवाको परिमाण',                'Float',    130),
			('uom',                   None,    'वस्तु वा सेवाको एकाइ',                                            'Data',     100),
			('purchase',              None,    'जम्मा फिर्ता मूल्य (रु)',                                        'Currency', 140),
			('tax_free_purchase',     None,    'कर छुट हुने वस्तु वा सेवाको फिर्ता मूल्य (रु)',                   'Currency', 160),
			('taxable_purchase',      'txp',   'मूल्य (रु)',                                                     'Currency', 120),
			('vat',                   'txp',   'कर (रु)',                                                        'Currency', 110),
			('taxable_import',        'tim',   'मूल्य (रु)',                                                     'Currency', 120),
			('import_vat',            'tim',   'कर (रु)',                                                        'Currency', 110),
			('capitalized_purchase',  'cap',   'मूल्य (रु)',                                                     'Currency', 120),
			('capitalized_vat',       'cap',   'कर (रु)',                                                        'Currency', 110),
		]
	return [
		('miti',                  None,    'मिति',                                                'Data',     120),
		('voucher_no',            'bijak', 'बीजक नं.',                                             'Data',     170),
		('pragyanpatra_no',       'bijak', 'प्रज्ञापनपत्र नं.',                                    'Data',     130),
		('supplier_name',         'bijak', 'आपूर्तिकर्ताको नाम',                                   'Data',     180),
		('vat_number',            'bijak', 'आपूर्तिकर्ताको स्थायी लेखा नम्बर',                      'Data',     150),
		('purchase',              None,    'जम्मा खरिद मूल्य (रु)',                                'Currency', 140),
		('tax_free_purchase',     None,    'कर छुट हुने वस्तुवा सेवाको खरिद / पैठारी मूल्य (रु)',  'Currency', 160),
		('taxable_purchase',      'txp',   'मूल्य (रु)',                                           'Currency', 120),
		('vat',                   'txp',   'कर (रु)',                                              'Currency', 110),
		('taxable_import',        'tim',   'मूल्य (रु)',                                           'Currency', 120),
		('import_vat',            'tim',   'कर (रु)',                                              'Currency', 110),
		('capitalized_purchase',  'cap',   'मूल्य (रु)',                                           'Currency', 120),
		('capitalized_vat',       'cap',   'कर (रु)',                                              'Currency', 110),
	]


def get_govt_columns(is_return=False):
	"""Column defs (Nepali labels) for the on-screen/print govt view — Purchase Book (default)
	or Purchase Return Book (is_return)."""
	return [
		{"fieldname": f, "label": label, "fieldtype": ftype, "width": width}
		for f, _group, label, ftype, width in _govt_fields(is_return)
	]


def _build_govt_layout(selected_columns, is_return=False):
	"""Row1 (group super-headers only) for the govt book layout. Individual (non-grouped)
	sub labels come from get_govt_columns() as the real column headers, not from this overlay."""
	show_all = not selected_columns
	visible = lambda f: show_all or f in selected_columns
	groups = _govt_groups(is_return)

	body_fields = [
		{'key': f, 'group': group, 'sub': label, 'css': 'l' if ftype == 'Data' else 'r', 'kind': 'currency' if ftype == 'Currency' else ('voucher' if f == 'voucher_no' else 'text')}
		for f, group, label, ftype, width in _govt_fields(is_return)
		if visible(f)
	]

	row1 = []
	i = 0
	while i < len(body_fields):
		bf = body_fields[i]
		if bf['group']:
			j = i
			while j < len(body_fields) and body_fields[j]['group'] == bf['group']:
				j += 1
			# Merged group headers are always centered over their span, regardless of
			# whether the underlying columns are left (Data) or right (Currency) aligned.
			row1.append({'text': groups[bf['group']], 'css': 'c', 'colspan': j - i, 'rowspan': 1})
			i = j
		else:
			row1.append({'text': bf['sub'], 'css': bf['css'], 'colspan': 1, 'rowspan': 2})
			i += 1

	return {'row1': row1, 'body_fields': body_fields}


@frappe.whitelist()
def get_govt_header_html(selected_columns=None, is_return=None):
	"""Row1 (Nepali group super-headers only) as an HTML fragment, for the on-screen report
	to show above its own (already Nepali-labelled) column headers. Standalone (non-merged)
	columns render blank here — their label is the real column header underneath (row 2)."""
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	layout = _build_govt_layout(selected_columns, is_return=bool(frappe.utils.cint(is_return)))

	tds = []
	for c in layout['row1']:
		is_group_cell = c.get('colspan', 1) > 1
		attrs = f' colspan="{c["colspan"]}"' if is_group_cell else ''
		text = c['text'] if is_group_cell else ''
		tds.append(f'<th class="{c["css"]}"{attrs}>{frappe.utils.escape_html(text)}</th>')

	return '<table><tr>' + ''.join(tds) + '</tr></table>'


@frappe.whitelist()
def get_print_html(filters, selected_columns=None, orientation=None):
	if isinstance(filters, str):
		filters = frappe._dict(json.loads(filters))
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	is_return = bool(filters.get('is_return'))
	_, data = execute(filters)
	supplier_display, ptype_display = _get_display_names(filters)
	govt_layout = _build_govt_layout(selected_columns, is_return=is_return)
	govt_heading = _get_govt_heading_info(filters)

	template_path = os.path.join(os.path.dirname(__file__), 'purchase_register_report_pdf.html')
	with open(template_path) as f:
		template_content = f.read()

	return frappe.render_template(template_content, {
		'filters': filters,
		'data': data,
		'fmt': _fmt_inr,
		'sc': selected_columns or [],
		'is_html_view': True,
		'orientation': orientation or 'Landscape',
		'supplier_display': supplier_display,
		'ptype_display': ptype_display,
		'govt_layout': govt_layout,
		'govt_heading': govt_heading,
	})


@frappe.whitelist()
def download_pdf(filters, orientation=None, selected_columns=None, view=None):
	from frappe.utils.pdf import get_pdf

	if isinstance(filters, str):
		filters = frappe._dict(json.loads(filters))
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	orientation = orientation if orientation in ('Portrait', 'Landscape') else 'Landscape'

	is_return = bool(filters.get('is_return'))
	_, data = execute(filters)
	supplier_display, ptype_display = _get_display_names(filters)
	govt_layout = _build_govt_layout(selected_columns, is_return=is_return)
	govt_heading = _get_govt_heading_info(filters)

	template_path = os.path.join(os.path.dirname(__file__), 'purchase_register_report_pdf.html')
	with open(template_path) as f:
		template_content = f.read()

	# sc = columns currently shown in the report (picked in the print dialog, or whatever
	# is left after removing columns in the datatable). Empty => show all.
	# Page numbers are rendered in the template body (manual pagination) so they work on
	# plain/unpatched wkhtmltopdf — no --footer-html (needs patched-Qt, silently ignored).
	html = frappe.render_template(template_content, {
		'filters': filters,
		'supplier_display': supplier_display,
		'ptype_display': ptype_display,
		'data': data,
		'fmt': _fmt_inr,
		'sc': selected_columns or [],
		'orientation': orientation,
		'govt_layout': govt_layout,
		'govt_heading': govt_heading,
	})

	options = {
		'page-size': 'A4',
		'orientation': orientation,
		'margin-top': '10mm',
		'margin-right': '8mm',
		'margin-bottom': '15mm',
		'margin-left': '8mm',
		'encoding': 'UTF-8',
		'enable-local-file-access': None,
	}
	pdf_data = get_pdf(html, options)

	frappe.response.filename = 'purchase_register.pdf'
	frappe.response.filecontent = pdf_data
	# view=1 (Print) → open inline in the browser tab; otherwise download the file.
	frappe.response.type = 'pdf' if frappe.utils.cint(view) else 'download'


@frappe.whitelist()
def download_excel(filters, selected_columns=None):
	"""XLSX with the same govt VAT Purchase Book heading/column-group layout as the PDF
	(title, PAN/नाम/साल line, merged group headers) — the stock Export menu can't produce
	this (flat columns + optional filter list only), so this is a dedicated button."""
	from io import BytesIO

	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, Side, Border
	from openpyxl.utils import get_column_letter
	from frappe.desk.utils import provide_binary_file

	if isinstance(filters, str):
		filters = frappe._dict(json.loads(filters))
	if isinstance(selected_columns, str):
		selected_columns = json.loads(selected_columns)

	_, data = execute(filters)
	is_return = bool(filters.get('is_return'))

	wb = Workbook()
	ws = wb.active
	ws.title = 'Purchase Register'

	thin = Side(style='thin', color='999999')
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	center = Alignment(horizontal='center', vertical='center', wrap_text=True)
	left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
	right_align = Alignment(horizontal='right', vertical='center')

	layout = _build_govt_layout(selected_columns, is_return=is_return)
	groups = _govt_groups(is_return)
	heading = _get_govt_heading_info(filters)
	body_fields = layout['body_fields']
	total_cols = len(body_fields)

	title = 'खरिद फिर्ता खाता' if is_return else 'खरिद खाता'
	c = ws.cell(row=1, column=1, value=title)
	c.font = Font(bold=True, size=16)
	c.alignment = center
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

	c = ws.cell(row=2, column=1, value='(नियम २३ को उपनियम (१) को खण्ड  (छ) संग सम्बन्धित )')
	c.alignment = center
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

	pan_line = (
		f"करदाता दर्ता नं (PAN) : {heading['pan']}      "
		f"करदाताको नाम: {heading['company_name']}      "
		f"साल: {heading['bs_year']}      "
		f"कर अवधि: {heading['tax_period']}"
	)
	c = ws.cell(row=4, column=1, value=pan_line)
	c.font = Font(bold=True)
	c.alignment = left_align
	ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)

	group_row, sub_row = 5, 6
	i, col_idx = 0, 1
	while i < len(body_fields):
		bf = body_fields[i]
		if bf['group']:
			j = i
			while j < len(body_fields) and body_fields[j]['group'] == bf['group']:
				j += 1
			span = j - i
			c = ws.cell(row=group_row, column=col_idx, value=groups[bf['group']])
			c.font, c.alignment, c.border = Font(bold=True), center, border
			ws.merge_cells(start_row=group_row, start_column=col_idx, end_row=group_row, end_column=col_idx + span - 1)
			for k in range(span):
				sc = ws.cell(row=sub_row, column=col_idx + k, value=body_fields[i + k]['sub'])
				sc.font, sc.alignment, sc.border = Font(bold=True), center, border
			i, col_idx = j, col_idx + span
		else:
			c = ws.cell(row=group_row, column=col_idx, value=bf['sub'])
			c.font, c.alignment, c.border = Font(bold=True), center, border
			ws.merge_cells(start_row=group_row, start_column=col_idx, end_row=sub_row, end_column=col_idx)
			i, col_idx = i + 1, col_idx + 1

	row = sub_row + 1
	for d in data:
		for ci, bf in enumerate(body_fields, start=1):
			val = d.get(bf['key'])
			if bf['kind'] == 'voucher':
				val = (val or '').split('::')[0]
			c = ws.cell(row=row, column=ci, value=val)
			c.border = border
			if bf['kind'] == 'currency':
				c.alignment = right_align
				c.number_format = '#,##0.00'
			else:
				c.alignment = left_align
			if d.get('bold'):
				c.font = Font(bold=True)
		row += 1

	for i in range(1, total_cols + 1):
		ws.column_dimensions[get_column_letter(i)].width = 18

	buf = BytesIO()
	wb.save(buf)
	provide_binary_file('purchase_register', 'xlsx', buf.getvalue())


def execute(filters=None):
	filters = filters or {}
	is_return = bool(filters.get("is_return"))
	# Always the official govt VAT book view now — "खरिद खाता" (Purchase) when Is Return is
	# unticked, "खरिद फिर्ता खाता" (Purchase Return) when ticked. Only fields with a govt
	# Nepali equivalent are shown; there is no more plain flat register.
	columns = get_govt_columns(is_return=is_return)
	data = get_data(filters)

	# Purchase Returns store amounts as negatives in the DB. Show them as positives in the register.
	if is_return and data:
		numeric_fields = [c["fieldname"] for c in columns if c.get("fieldtype") in ("Currency", "Float")]
		for row in data:
			for f in numeric_fields:
				if row.get(f) is not None:
					row[f] = abs(row[f])

	if data:
		total = {"voucher_no": _("Total"), "bold": 1}
		for col in columns:
			if col.get("fieldtype") in ("Currency", "Float"):
				total[col["fieldname"]] = sum(row.get(col["fieldname"]) or 0 for row in data)
		data.append(total)
	return columns, data


def get_data(filters):
	is_return = 1 if filters.get("is_return") else 0
	conditions = f"pi.docstatus = 1 AND pi.is_return = {is_return}"

	if filters.get("company"):
		placeholders = ", ".join(["'{}'".format(c) for c in filters.get("company")])
		conditions += " AND pi.company IN ({})".format(placeholders)
	if filters.get("from_date"):
		conditions += " AND pi.posting_date >= %(from_date)s"
	if filters.get("to_date"):
		conditions += " AND pi.posting_date <= %(to_date)s"
	if filters.get("supplier"):
		placeholders = ", ".join(["'{}'".format(c) for c in filters.get("supplier")])
		conditions += " AND pi.supplier IN ({})".format(placeholders)
	if filters.get("purchase_type"):
		placeholders = ", ".join(["'{}'".format(c) for c in filters.get("purchase_type")])
		conditions += " AND pi.custom_purchase_type IN ({})".format(placeholders)

	rows = frappe.db.sql(
		f"""
		SELECT
			pi.posting_date                                                                          AS date,
			SUBSTRING_INDEX(pi.custom_nepali_miti, ' ', 1)                                           AS miti,
			pi.custom_purchase_type                                                                  AS purchase_type,
			CONCAT(IFNULL(pi.custom_name, pi.name), '::', pi.name)                                  AS voucher_no,
			pi.bill_no                                                                               AS supplier_invoice_no,
			pi.bill_date                                                                             AS supplier_invoice_date,
			SUBSTRING_INDEX(pi.custom_supplier_invoice_miti, ' ', 1)                                 AS supplier_invoice_miti,
			pi.supplier_name                                                                         AS supplier_name,
			s.tax_id                                                                                 AS vat_number,
			pi.custom_total_amount_including_excise                                                  AS purchase,
			SUM(CASE WHEN pii.custom_vat_apply_on = 'VAT 0%%'                                                                          THEN pii.amount ELSE 0 END) AS tax_free_purchase,
			SUM(CASE WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 0 AND s.custom_territory = 'Nepal'    THEN pii.amount ELSE 0 END) AS taxable_purchase,
			SUM(CASE WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 0 AND s.custom_territory = 'Nepal'    THEN pii.custom_vat_amount ELSE 0 END) AS vat,
			SUM(CASE WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 0 AND s.custom_territory != 'Nepal'   THEN pii.amount ELSE 0 END) AS taxable_import,
			SUM(CASE WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 0 AND s.custom_territory != 'Nepal'   THEN pii.custom_vat_amount ELSE 0 END) AS import_vat,
			SUM(CASE WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 1 AND s.custom_territory = 'Nepal'    THEN pii.amount ELSE 0 END) AS capitalized_purchase,
			SUM(CASE WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 1 AND s.custom_territory = 'Nepal'    THEN pii.custom_vat_amount ELSE 0 END) AS capitalized_vat,
			SUM(CASE WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 0 AND s.custom_territory = 'Nepal'    THEN pii.custom_vat_amount
			         WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 0 AND s.custom_territory != 'Nepal'   THEN pii.custom_vat_amount
			         WHEN pii.custom_vat_apply_on IN ('VAT 13%%','Amount') AND i.is_fixed_asset = 1 AND s.custom_territory = 'Nepal'    THEN pii.custom_vat_amount
			         ELSE 0 END)                                                                      AS total_vat,
			SUM(pii.qty)                                                                             AS qty,
			GROUP_CONCAT(DISTINCT i.item_name SEPARATOR ', ')                                        AS item_description,
			GROUP_CONCAT(DISTINCT pii.uom SEPARATOR ', ')                                            AS uom
		FROM `tabPurchase Invoice` pi
		JOIN `tabPurchase Invoice Item` pii ON pii.parent = pi.name
		JOIN `tabSupplier` s ON s.name = pi.supplier
		JOIN `tabItem` i ON i.name = pii.item_code
		WHERE {conditions}
		GROUP BY pi.name
		ORDER BY pi.posting_date ASC
		""",
		filters,
		as_dict=True,
	)
	return rows
