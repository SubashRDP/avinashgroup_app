# Copyright (c) 2026, Raindrop and contributors
# For license information, please see license.txt

"""Read the IRD monthly sales book (बिक्री खाता) the old software exported.

A SECOND legacy format, and not a variant of the annexure — a different report
altogether. NGI was handed over as one 21-column English "Annexure-7" sheet per
year; NGK comes as twelve Nepali monthly sheets per year, headed "IRD Sales Data
Sync Report", under Rule 23(1). Same purpose, different columns, so it needs its
own reader (see import_legacy_annexure.py for the annexure one).

What it has that the annexure lacked: the exempt-sales column, so the
taxable/exempt split is real here instead of forced to zero.

What it lacks: every status column. No Sync with IRD, no Printed, no Printed By,
no Realtime. Those come from elsewhere — the operator's instruction of
2026-08-06 is that these rows came back FROM the IRD, so sync is true by
construction; print history comes from the print registers; and realtime is
recorded false because nothing here can tell us whether it was.

Sales and returns share the sheet, separated by the प्रकार (type) column:
"Sales" -> CBMS Bill, "SalesReturn" -> CBMS Bill Return. They are kept apart
here rather than left for the caller to sort out.

Read-only. Nothing in this module writes.
"""

import collections
import datetime
import glob
import json
import os
import re

import frappe
from frappe.utils import flt

FIRST_DATA_ROW = 7

# 0-based positions in the monthly sheet.
COL = {
    "type": 0,          # प्रकार            Sales / SalesReturn
    "date_bs": 1,       # बीजक मिति         2082.04.01
    "number": 2,        # बीजक नम्बर        NGK000002/82-83
    "buyer_name": 3,    # खरिदकर्ताको नाम
    "pan": 4,           # स्थायी लेखा नम्बर
    "grand_total": 5,   # जम्मा बिक्री       tax included
    "exempt": 6,        # कर छुटको बिक्री
    "taxable": 7,       # करयोग्य बिक्री
    "vat": 8,           # कर
    "export": 9,        # निकासी
    "entered_at": 10,   # प्रविस्ट मिति      AD, M/D/YYYY h:mm:ss AM
}

SALE, RETURN = "Sales", "SalesReturn"

# The monthly sheets travel with the code, as the annexure sheets do. Kept in a
# subfolder so the annexure importer's own non-recursive glob over sheets/ does
# not pick them up — the two formats must not be fed to each other's reader.
DEFAULT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sheets", "ngk")


def _open(path):
    """Workbook rows, with openpyxl's colour validator relaxed for the read.

    These exports carry a colour openpyxl rejects outright ("Colors must be aRGB
    hex values"), which otherwise stops the file being opened at all. Restored
    in a finally so nothing else in the process is affected.
    """
    import openpyxl.styles.colors as colors

    original = colors.RGB.__set__
    colors.RGB.__set__ = lambda self, instance, value: instance.__dict__.__setitem__(
        self.name, value
    )
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        rows = [
            row
            for row in wb.active.iter_rows(min_row=FIRST_DATA_ROW, values_only=True)
            if len(row) > COL["number"] and row[COL["number"]]
        ]
        wb.close()
        return rows
    finally:
        colors.RGB.__set__ = original


def _entered_at(value):
    """"4/15/2026 10:31:12 AM" -> datetime, or None. Already AD, not BS."""
    if not value:
        return None
    if isinstance(value, datetime.datetime):
        return value
    text = str(value).strip()
    for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _fiscal_year(number, folder_hint=None):
    """"NGK000002/82-83" -> "82.83". Falls back to the folder name ("NGK-82.83")."""
    _, _, tail = str(number or "").partition("/")
    bits = tail.split("-")
    if len(bits) == 2 and all(len(b) == 2 and b.isdigit() for b in bits):
        return f"{bits[0]}.{bits[1]}"
    if folder_hint:
        m = re.search(r"(\d{2})[.\-](\d{2})", folder_hint)
        if m:
            return f"{m.group(1)}.{m.group(2)}"
    return None


def read_sales_book(paths):
    """Normalised rows from one or more monthly sheets.

    Returns (sales, returns, info). Duplicates are dropped: the FY 80/81 export
    repeats 38 credit notes verbatim inside a single month's file — same number,
    date and amount twice over — which would double-count them. A repeat that
    disagrees on amount is NOT a duplicate and is kept, with a count reported,
    because that is a different problem and should not be silently merged.
    """
    sales, returns = {}, {}
    exact_dupes = conflicting = skipped_unknown_type = 0

    for path in paths:
        hint = os.path.basename(os.path.dirname(path)) + os.path.basename(path)
        for raw in _open(path):
            row_type = str(raw[COL["type"]] or "").strip()
            if row_type not in (SALE, RETURN):
                skipped_unknown_type += 1
                continue

            number = str(raw[COL["number"]]).strip()
            grand_total = flt(raw[COL["grand_total"]])
            exempt = flt(raw[COL["exempt"]])
            row = {
                "number": number,
                "is_return": row_type == RETURN,
                "fiscal_year": _fiscal_year(number, hint),
                # the sheet writes 2082.04.01; bills store 2082-04-01
                "date_bs": str(raw[COL["date_bs"]] or "").strip().replace(".", "-"),
                "buyer_name": str(raw[COL["buyer_name"]] or "").strip(),
                "pan": str(raw[COL["pan"]] or "").strip(),
                "grand_total": grand_total,
                "exempt": exempt,
                "taxable": flt(raw[COL["taxable"]]),
                "vat": flt(raw[COL["vat"]]),
                "export": flt(raw[COL["export"]]),
                # CBMS total_sales excludes exempt; the report adds it back, so
                # Total Amount comes out as the sheet's जम्मा बिक्री exactly.
                "total_sales": grand_total - exempt,
                "entered_at": _entered_at(raw[COL["entered_at"]]),
                "source_file": os.path.basename(path),
            }

            bucket = returns if row["is_return"] else sales
            existing = bucket.get(number)
            if existing:
                if (
                    abs(existing["grand_total"] - row["grand_total"]) < 0.01
                    and existing["date_bs"] == row["date_bs"]
                ):
                    exact_dupes += 1
                else:
                    conflicting += 1
                continue
            bucket[number] = row

    info = {
        "files_read": len(paths),
        "exact_duplicates_dropped": exact_dupes,
        "same_number_different_amount": conflicting,
        "rows_of_unknown_type_skipped": skipped_unknown_type,
    }
    return list(sales.values()), list(returns.values()), info


def sheets_in(folder):
    """Every monthly sheet under a folder, recursively."""
    return sorted(glob.glob(os.path.join(folder, "**", "*.xlsx"), recursive=True))


def run(company, folder=DEFAULT_FOLDER, fiscal_year=None, commit=False, limit=None):
    """Create CBMS Bills and Bill Returns from the monthly sales book.

    Field decisions are the operator's, 2026-08-06:

      sync_status  Synced, is_synced 1 — these rows came back FROM the IRD, so
                   they were reported by definition. This is also what keeps
                   them away from retry_failed_cbms_syncs; belt and braces now
                   that send_bill_to_cbms refuses anything below the company's
                   Send From Date.
      is_realtime  0 — the sheet cannot tell us, and false is the honest
                   default rather than a guess.
      exempt split real, from the sheet's own कर छुटको बिक्री column.
      Print history comes from the registers, NOT from here — see
      import_legacy_print_counts.run_print_log_backfill.

    A return needs the ORIGINAL invoice's number, which the sheet does not
    carry; it is read from the credit note's return_against in ERPNext, the same
    way create_cbms_bill_return does it. A credit note with no return_against
    cannot name its original and is skipped rather than guessed at.

    DRY RUN unless commit=True. Idempotent: an invoice that already has a CBMS
    record is skipped, so this resumes after an interruption.
    """
    from avinashgroup_app.legacy_annexure_import.import_legacy_annexure import (
        _entered_by,
        _fiscal_year_span,
        _resolve_invoices,
    )

    paths = sheets_in(folder)
    if not paths:
        frappe.throw(f"No .xlsx sheets found under {folder}")

    sales, returns, info = read_sales_book(paths)
    if fiscal_year:
        wanted = fiscal_year.replace("/", ".")
        sales = [r for r in sales if r["fiscal_year"] == wanted]
        returns = [r for r in returns if r["fiscal_year"] == wanted]

    seller_pan = frappe.get_cached_value("Company", company, "tax_id") or ""
    created = {"bills": 0, "returns": 0}
    skipped = {"already": 0, "not_in_erpnext": 0, "ambiguous": 0, "no_original": 0}
    samples = []

    for rows, is_return in ((sales, False), (returns, True)):
        by_year = collections.defaultdict(list)
        for r in rows:
            by_year[r["fiscal_year"]].append(r)

        for fy in sorted(by_year):
            span = _fiscal_year_span(fy)
            numbers = {r["number"] for r in by_year[fy]}
            mapping, ambiguous = _resolve_invoices(
                numbers,
                company,
                from_date=span.year_start_date if span else None,
                to_date=span.year_end_date if span else None,
            )

            doctype = "CBMS Bill Return" if is_return else "CBMS Bill"
            existing = set()
            names = [inv.name for inv in mapping.values()]
            for chunk in (names[i : i + 5000] for i in range(0, len(names), 5000)):
                existing.update(
                    frappe.get_all(
                        doctype, filters={"sales_invoice": ["in", chunk]}, pluck="sales_invoice"
                    )
                )

            for row in by_year[fy]:
                invoice = mapping.get(row["number"])
                if not invoice:
                    skipped["ambiguous" if row["number"] in ambiguous else "not_in_erpnext"] += 1
                    continue
                if invoice.name in existing:
                    skipped["already"] += 1
                    continue
                if limit and (created["bills"] + created["returns"]) >= int(limit):
                    break

                fields = {
                    "doctype": doctype,
                    "company": invoice.company,
                    "sales_invoice": invoice.name,
                    "fiscal_year": fy,
                    "created_by": _entered_by(invoice),
                    "buyer_name": row["buyer_name"],
                    "buyer_pan": row["pan"],
                    "seller_pan": seller_pan,
                    "total_sales": row["total_sales"],
                    "tax_exempted_sales": row["exempt"],
                    "taxable_sales_vat": row["taxable"],
                    "vat": row["vat"],
                    "export_sales": row["export"],
                    "discount": 0,
                    "sync_status": "Synced",
                    "is_synced": 1,
                    "is_realtime": 0,
                    "last_attempt": row["entered_at"],
                    "datetime_client": row["entered_at"] or invoice.posting_date,
                }

                if is_return:
                    original = _original_number(invoice.name)
                    if not original:
                        skipped["no_original"] += 1
                        continue
                    fields.update(
                        {
                            "ref_invoice_number": original,
                            "credit_note_number": row["number"],
                            "credit_note_date": invoice.posting_date,
                            "credit_note_date_bs": row["date_bs"],
                            "reason_for_return": "Goods Returned",
                        }
                    )
                else:
                    fields.update(
                        {
                            "invoice_number": row["number"],
                            "invoice_date": invoice.posting_date,
                            "invoice_date_bs": row["date_bs"],
                        }
                    )

                if commit:
                    frappe.get_doc(fields).insert(ignore_permissions=True)
                created["returns" if is_return else "bills"] += 1

                if len(samples) < 4:
                    samples.append(
                        {
                            "doctype": doctype,
                            "invoice": invoice.name,
                            "number": row["number"],
                            "bs_date": row["date_bs"],
                            "total_sales": row["total_sales"],
                            "exempt": row["exempt"],
                            "vat": row["vat"],
                            "created_by": fields["created_by"],
                            "entered_at": str(row["entered_at"]),
                        }
                    )

    if commit:
        frappe.db.commit()

    result = {
        "dry_run": not commit,
        "folder": folder,
        "company": company,
        **info,
        ("created" if commit else "would_create"): created,
        "skipped": skipped,
        "samples": samples,
        "note": "Print history is NOT written here — run run_print_log_backfill "
        "over the NGK registers for that.",
    }
    print(json.dumps(result, indent=1, default=str))
    return result


def _original_number(credit_note):
    """The branch number of the invoice a credit note was raised against."""
    against = frappe.db.get_value("Sales Invoice", credit_note, "return_against")
    if not against:
        return None
    return frappe.db.get_value("Sales Invoice", against, "custom_branch_name") or against


def match(company, folder=DEFAULT_FOLDER, fiscal_year=None, samples=10):
    """Match the sales book against Sales Invoice — and write nothing.

    The question this answers, before any import is attempted: how many of the
    sheet's invoices actually exist in ERPNext, and which do not. Matching is on
    custom_branch_name within COMPANY and FISCAL YEAR, the scope a legacy number
    is unique in — the numbering engine works the same way.

      bench --site <site> execute \\
        avinashgroup_app.legacy_annexure_import.sales_book.match \\
        --kwargs "{'company': 'Nepal Gas Udhyog (Karnali) Pvt. Ltd.'}"
    """
    from avinashgroup_app.legacy_annexure_import.import_legacy_annexure import (
        _fiscal_year_span,
        _resolve_invoices,
    )

    paths = sheets_in(folder)
    if not paths:
        frappe.throw(f"No .xlsx sheets found under {folder}")

    sales, returns, info = read_sales_book(paths)
    if fiscal_year:
        wanted = fiscal_year.replace("/", ".")
        sales = [r for r in sales if r["fiscal_year"] == wanted]
        returns = [r for r in returns if r["fiscal_year"] == wanted]

    per_year = collections.OrderedDict()
    for years_rows, label in ((sales, "sales"), (returns, "returns")):
        by_year = collections.defaultdict(list)
        for r in years_rows:
            by_year[r["fiscal_year"]].append(r)

        for fy in sorted(by_year):
            span = _fiscal_year_span(fy)
            numbers = {r["number"] for r in by_year[fy]}
            mapping, ambiguous = _resolve_invoices(
                numbers,
                company,
                from_date=span.year_start_date if span else None,
                to_date=span.year_end_date if span else None,
            )
            missing = sorted(numbers - set(mapping))
            entry = per_year.setdefault(fy, {})
            entry[label] = {
                "sheet_rows": len(by_year[fy]),
                "matched": len(numbers) - len(missing),
                "not_in_erpnext": len(missing),
                "ambiguous": len(ambiguous),
                "missing_sample": missing[:samples],
                "fiscal_year_record": bool(span),
            }

    result = {
        "folder": folder,
        "company": company,
        **info,
        "sales_rows": len(sales),
        "return_rows": len(returns),
        "per_year": per_year,
    }
    print(json.dumps(result, indent=1, default=str))
    return result
