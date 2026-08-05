# Copyright (c) 2026, Raindrop and contributors
# For license information, please see license.txt

"""Load the old billing software's VAT Annexure 7 exports into CBMS Bill.

WHY THIS EXISTS
The Materialized Report draws its rows FROM CBMS Bill, and a bill is only ever
written when an invoice is submitted through the CBMS hook — which began at
e-billing go-live. Every invoice before that has no bill, so the report is empty
for 79/80-82/83 no matter what the invoices themselves contain. The old software
exported those years as "Annexure-7" sheets; this loads them back, so the report
reproduces them with no change to the report itself.

WHAT IT READS
The annexure sheet: header on row 6, data from row 7, 21 columns. Column B is
the old invoice number, which is matched to a Sales Invoice through
custom_branch_name. Rows whose invoice is not in ERPNext are SKIPPED and
counted — the 79/80 sheet in particular names invoices that were never migrated.

WHAT IT WRITES
One CBMS Bill per matched row, carrying the sheet's amounts and IRD status, plus
one Sales Invoice Print Log row per row the sheet marks Printed. Nothing is sent
anywhere: this only records what the old software already did.

  Entered By  -> created_by, taken from the INVOICE's custom_created_by, not
                 from the sheet. The sheet names the old software's user
                 ("ASHISH"), which is not a Frappe user; the audit field names
                 the clerk. Decision of 2026-08-05.
  Printed *   -> a Print Log row with the sheet's Printed By verbatim and its
                 Printed Time (BS) converted to AD as `creation`.

  !!! READ BEFORE RUNNING ON LIVE !!!
  scheduler.retry_failed_cbms_syncs re-sends every CBMS Bill whose sync_status
  is not "Synced", for any company with CBMS enabled. Reporting a bill to IRD
  cannot be undone. The sheets mark a handful of rows unsynced (85 across the
  four years), and importing those as Failed puts them in the retry job's path.
  So: either run with unsynced_status="Synced" (the default — nothing becomes
  eligible for a send), or disable the retry job for the duration and decide
  what those rows should be. The default is chosen so that a careless run cannot
  transmit anything.

DRY RUN unless commit=True. Idempotent: an invoice that already has a CBMS Bill
is skipped whole, so a second run writes nothing.

Usage (from the bench directory):

    bench --site <site> execute \\
      avinashgroup_app.legacy_annexure_import.import_legacy_annexure.run \\
      --kwargs "{'path': '/home/sijan/Downloads/Materialized Report_NGI 82.83.xlsx'}"

    bench --site <site> execute \\
      avinashgroup_app.legacy_annexure_import.import_legacy_annexure.run \\
      --kwargs "{'path': '...', 'commit': True}"
"""

import json
import os

import frappe
import nepali_datetime
from frappe.utils import flt

HEADER_ROW = 6
FIRST_DATA_ROW = 7

# 0-based positions in the 21-column annexure sheet.
COL = {
    "fiscal_year": 0,
    "invoice_number": 1,
    "invoice_date": 2,
    "customer_name": 3,
    "pan": 4,
    "amount": 5,
    "discount": 6,
    "taxable": 7,
    "tax": 8,
    "total": 9,
    "synced": 10,
    "printed": 11,
    "active": 12,
    "printed_time": 13,
    "entered_by": 14,
    "printed_by": 15,
    "realtime": 16,
    "synced_at": 20,
}

PRINT_LOG_FIELDS = (
    "name", "creation", "modified", "modified_by", "owner", "docstatus", "idx",
    "sales_invoice", "customer", "customer_name", "branch_name", "company",
    "copy_number", "printed_by",
)


def _load_sheet(path):
    """Rows of the annexure sheet, from the first data row down.

    The exports carry a colour their own writer produced but openpyxl rejects
    ("Colors must be aRGB hex values"), which stops the file being opened at
    all. The validator is relaxed for this read only — nothing here looks at
    formatting.
    """
    import openpyxl.styles.colors as colors

    original = colors.RGB.__set__
    colors.RGB.__set__ = lambda self, instance, value: instance.__dict__.__setitem__(
        self.name, value
    )
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [
            row
            for row in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True)
            if row[COL["invoice_number"]]
        ]
        wb.close()
        return rows
    finally:
        colors.RGB.__set__ = original


def _bs_to_ad_datetime(value):
    """"2082/04/01 10:53:33 AM" (BS) -> an AD datetime, or None.

    The sheet writes print times 12-hour and sync times 24-hour; both are
    accepted. A date that the BS calendar does not contain returns None rather
    than raising — one unparseable cell must not abandon 36,000 rows.
    """
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None

    date_part, _, time_part = text.partition(" ")
    bits = date_part.replace("-", "/").split("/")
    if len(bits) != 3:
        return None
    try:
        bs = nepali_datetime.date(int(bits[0]), int(bits[1]), int(bits[2]))
        ad = bs.to_datetime_date()
    except Exception:
        return None

    clock, meridiem = (time_part.rsplit(" ", 1) + [""])[:2] if time_part else ("", "")
    hh = mm = ss = 0
    if clock:
        parts = (clock.split(":") + ["0", "0", "0"])[:3]
        try:
            hh, mm, ss = (int(float(p)) for p in parts)
        except ValueError:
            hh = mm = ss = 0
    upper = meridiem.upper()
    if upper == "PM" and hh < 12:
        hh += 12
    elif upper == "AM" and hh == 12:
        hh = 0

    import datetime as _dt

    return _dt.datetime(ad.year, ad.month, ad.day, hh, mm, ss)


def _bs_date_stored(value):
    """Sheet's "2082/04/01" -> the "2082-04-01" form CBMS Bill stores."""
    if not value:
        return None
    return str(value).strip().replace("/", "-")


def _yes(value):
    return str(value or "").strip().lower() == "yes"


def _fiscal_year_dotted(rows, override=None):
    """Bill fiscal_year ("82.83") for the sheet.

    Derived from the invoice numbers, which carry the year ("NGI000004/82-83"),
    NOT from the sheet's Fiscal Year column — that prints a two-year span
    ("2082-84") which is not a Fiscal Year record.
    """
    if override:
        return override.replace("/", ".")
    for row in rows:
        number = str(row[COL["invoice_number"]] or "").strip()
        _, _, tail = number.partition("/")
        bits = tail.split("-")
        if len(bits) == 2 and all(len(b) == 2 and b.isdigit() for b in bits):
            return f"{bits[0]}.{bits[1]}"

    # 79/80 predates the year suffix — its numbers are a bare "NGI/030020". Fall
    # back to the sheet's own Fiscal Year column, which reads "2079-81": four
    # digits for the opening year, then the year TWO on (the legacy header's
    # two-year accounting window). The fiscal year is opening..opening+1.
    for row in rows:
        printed = str(row[COL["fiscal_year"]] or "").strip()
        head = printed.split("-")[0]
        if len(head) == 4 and head.isdigit():
            start = int(head) % 100
            return f"{start:02d}.{(start + 1) % 100:02d}"
    return None


def _fiscal_year_span(fy_dotted):
    """AD start/end of the Fiscal Year a bill's dotted year names ("82.83").

    Fiscal Year records are named with a slash. Returns None when the site has
    no record for that year, in which case the caller falls back to matching on
    the number and company alone.
    """
    if not fy_dotted:
        return None
    return frappe.db.get_value(
        "Fiscal Year",
        fy_dotted.replace(".", "/"),
        ["year_start_date", "year_end_date"],
        as_dict=True,
    )


def _resolve_invoices(numbers, company=None, from_date=None, to_date=None):
    """old invoice number (custom_branch_name) -> Sales Invoice row.

    A legacy number identifies an invoice only within a COMPANY and a FISCAL
    YEAR — the old ERPs numbered each company separately and restarted the
    series every year, which is the same scope the numbering engine works in
    (custom_code/Override/naming_series.py). NGK/000001 exists in both 77/78 and
    79/80; NGI000004/82-83 could repeat under another company.

    So the lookup is narrowed by company and by the sheet's fiscal year, applied
    as a posting-date window. Inside that scope the number is unique, and a
    number that still resolves to several invoices is reported rather than
    guessed at.
    """
    holders = {}
    filters = {"custom_branch_name": ["in", list(numbers)]}
    if company:
        filters["company"] = company
    if from_date and to_date:
        filters["posting_date"] = ["between", [from_date, to_date]]
    for r in frappe.get_all(
        "Sales Invoice",
        filters=filters,
        fields=[
            "name", "company", "custom_branch_name", "posting_date", "docstatus",
            "customer", "customer_name", "custom_created_by", "owner",
        ],
        limit_page_length=0,
    ):
        holders.setdefault(r.custom_branch_name, []).append(r)
    mapping = {k: v[0] for k, v in holders.items() if len(v) == 1}
    ambiguous = {k: [r.name for r in v] for k, v in holders.items() if len(v) > 1}
    return mapping, ambiguous


def _log_name(sales_invoice):
    """Print Log docname for an imported row: the invoice's own name with a
    suffix, e.g. "NGI-SB-79/80-00010-L1".

    Was frappe.generate_hash(length=10), which collided on live: ten hex
    characters over ~92,000 rows is a birthday collision waiting to happen, and
    it aborted the 81/82 load with "Duplicate entry ... for key 'PRIMARY'".
    Deriving the name from the invoice makes a collision impossible rather than
    unlikely, and makes a re-run land on the same names instead of new ones.

    One row per invoice is written by this importer (the sheet records a single
    print time), hence the fixed suffix. Doc names allow 140 characters; an
    invoice name is nowhere near that.
    """
    return f"{sales_invoice}-L1"


def _entered_by(invoice):
    """Clerk name for CBMS Bill.created_by — the invoice's audit field, resolved
    to a full name, with owner only as a fallback."""
    user = invoice.custom_created_by or invoice.owner
    if not user:
        return None
    return frappe.db.get_value("User", user, "full_name") or user


def run(path, company=None, fiscal_year=None, commit=False, limit=None,
        unsynced_status="Synced", with_print_log=True):
    """Import one annexure sheet. See the module docstring before running live."""
    if unsynced_status not in ("Synced", "Failed", "Pending"):
        frappe.throw("unsynced_status must be Synced, Failed or Pending")

    rows = _load_sheet(path)
    fy = _fiscal_year_dotted(rows, fiscal_year)
    if not fy:
        frappe.throw(f"Could not determine the fiscal year from {path}")

    numbers = {str(r[COL["invoice_number"]]).strip() for r in rows}
    span = _fiscal_year_span(fy)
    mapping, ambiguous = _resolve_invoices(
        numbers,
        company,
        from_date=span.year_start_date if span else None,
        to_date=span.year_end_date if span else None,
    )

    already = set()
    matched_names = [inv.name for inv in mapping.values()]
    for chunk in _chunks(matched_names, 5000):
        already.update(
            frappe.get_all(
                "CBMS Bill", filters={"sales_invoice": ["in", chunk]}, pluck="sales_invoice"
            )
        )

    seller_pans = {}
    bills, logs, samples = [], [], []
    skipped_unmatched, skipped_existing, skipped_ambiguous = [], 0, 0
    printed_rows = 0

    for row in rows:
        number = str(row[COL["invoice_number"]]).strip()
        invoice = mapping.get(number)
        if not invoice:
            if number in ambiguous:
                skipped_ambiguous += 1
            else:
                skipped_unmatched.append(number)
            continue
        if invoice.name in already:
            skipped_existing += 1
            continue
        if limit and len(bills) >= int(limit):
            break

        if invoice.company not in seller_pans:
            seller_pans[invoice.company] = (
                frappe.get_cached_value("Company", invoice.company, "tax_id") or ""
            )

        total = flt(row[COL["total"]])
        vat = flt(row[COL["tax"]])
        synced = _yes(row[COL["synced"]])

        bills.append(
            {
                "doctype": "CBMS Bill",
                "company": invoice.company,
                "sales_invoice": invoice.name,
                "invoice_number": number,
                "invoice_date": invoice.posting_date,
                "invoice_date_bs": _bs_date_stored(row[COL["invoice_date"]]),
                "fiscal_year": fy,
                "created_by": _entered_by(invoice),
                "buyer_name": row[COL["customer_name"]],
                "buyer_pan": str(row[COL["pan"]] or "").strip(),
                "seller_pan": seller_pans[invoice.company],
                # The sheet has no exempt/taxable split beyond these two columns,
                # so exempt stays 0 and total_sales carries the grand total. The
                # report computes Amount as total_sales + exempt - vat, which
                # reproduces the sheet's Amount column exactly.
                "total_sales": total,
                "tax_exempted_sales": 0,
                "taxable_sales_vat": flt(row[COL["taxable"]]),
                "vat": vat,
                "discount": flt(row[COL["discount"]]),
                "sync_status": "Synced" if synced else unsynced_status,
                "is_synced": 1 if synced else 0,
                "is_realtime": 1 if _yes(row[COL["realtime"]]) else 0,
                "last_attempt": _bs_to_ad_datetime(row[COL["synced_at"]]),
                "datetime_client": _bs_to_ad_datetime(row[COL["printed_time"]])
                or invoice.posting_date,
            }
        )

        if with_print_log and _yes(row[COL["printed"]]):
            printed_at = _bs_to_ad_datetime(row[COL["printed_time"]])
            if printed_at:
                printed_rows += 1
                logs.append(
                    (
                        _log_name(invoice.name),
                        printed_at,
                        frappe.utils.now(),
                        frappe.session.user,
                        frappe.session.user,
                        0,
                        0,
                        invoice.name,
                        invoice.customer,
                        invoice.customer_name,
                        number,
                        invoice.company,
                        1,
                        str(row[COL["printed_by"]] or "").strip() or None,
                    )
                )

        if len(samples) < 3:
            samples.append(
                {
                    "invoice": invoice.name,
                    "number": number,
                    "bs_date": bills[-1]["invoice_date_bs"],
                    "total_sales": bills[-1]["total_sales"],
                    "sync_status": bills[-1]["sync_status"],
                    "created_by": bills[-1]["created_by"],
                    "printed_at": str(logs[-1][1]) if logs and with_print_log else None,
                }
            )

    if commit:
        already_logged = set()
        for chunk in _chunks([b["sales_invoice"] for b in bills], 5000):
            already_logged.update(
                frappe.get_all(
                    "Sales Invoice Print Log",
                    filters={"sales_invoice": ["in", chunk]},
                    pluck="sales_invoice",
                )
            )
        for payload in bills:
            frappe.get_doc(payload).insert(ignore_permissions=True)
        fresh = [row for row in logs if row[7] not in already_logged]
        if fresh:
            frappe.db.bulk_insert("Sales Invoice Print Log", list(PRINT_LOG_FIELDS), fresh)
        frappe.db.commit()

    result = {
        "dry_run": not commit,
        "path": path,
        "fiscal_year": fy,
        "company": company or "(any)",
        "sheet_rows": len(rows),
        "invoices_matched": len(mapping),
        "bills_" + ("created" if commit else "to_create"): len(bills),
        "print_log_rows_" + ("created" if commit else "to_create"): len(logs),
        "rows_marked_printed": printed_rows,
        "skipped_already_have_a_bill": skipped_existing,
        "skipped_invoice_not_in_erpnext": len(skipped_unmatched),
        "skipped_unmatched_sample": sorted(skipped_unmatched)[:20],
        "skipped_ambiguous_number": skipped_ambiguous,
        "unsynced_status_used": unsynced_status,
        "samples": samples,
    }
    if unsynced_status != "Synced":
        result["retry_warning"] = (
            "Bills written as {0} are picked up by retry_failed_cbms_syncs and WILL be "
            "sent to the IRD. Disable that job before committing.".format(unsynced_status)
        )
    print(json.dumps(result, indent=1, default=str))
    return result


def _chunks(seq, size):
    seq = list(seq)
    for i in range(0, len(seq), size):
        yield seq[i : i + size]


# The sheets travel with the code, in sheets/ next to this file — the same way
# legacy_print_import ships its registers. A one-off historical load has no
# value without the exact file it was run from, and a path into somebody's
# Downloads folder does not survive the trip to a server.
DEFAULT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sheets")
SHEET_GLOB = "*.xlsx"


def _peek_fiscal_year(path):
    """Fiscal year of a sheet without reading all 36,000 rows.

    Streams the first data rows only and stops at the first invoice number that
    carries a year. Used to spot the same year exported twice — the folder holds
    "82.83", "82.83(1)" and "82.83(2)", which are byte-for-byte the same export.
    """
    import openpyxl.styles.colors as colors

    original = colors.RGB.__set__
    colors.RGB.__set__ = lambda self, instance, value: instance.__dict__.__setitem__(
        self.name, value
    )
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        head = []
        for i, row in enumerate(ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True)):
            head.append(row)
            if i >= 50:
                break
        wb.close()
        return _fiscal_year_dotted([r for r in head if r[COL["invoice_number"]]])
    except Exception:
        return None
    finally:
        colors.RGB.__set__ = original


def run_all(folder=DEFAULT_FOLDER, paths=None, company=None, commit=False,
            unsynced_status="Synced", with_print_log=True):
    """Import every annexure sheet in one go — the whole history, one command.

    Picks up `Materialized Report_*.xlsx` from `folder` (or an explicit `paths`
    list), keeps ONE file per fiscal year, and runs them oldest year first so
    the earliest invoices are in place before the later ones.

    Same rules as run(): DRY RUN unless commit=True, and an invoice that already
    has a CBMS Bill is skipped, so this is safe to repeat and safe to resume
    after an interruption.

      bench --site <site> execute \\
        avinashgroup_app.legacy_annexure_import.import_legacy_annexure.run_all \\
        --kwargs "{'company': 'Nepal Gas Udhyog Pvt. Ltd.', 'commit': True}"
    """
    import glob

    if paths is None:
        paths = sorted(glob.glob(os.path.join(folder, SHEET_GLOB)))
    if not paths:
        frappe.throw(f"No annexure sheets found in {folder}")

    # One file per fiscal year. Where a year appears more than once the plain
    # name wins over a "(1)" copy — same export, downloaded twice.
    chosen = {}
    duplicates = []
    for path in paths:
        fy = _peek_fiscal_year(path)
        if not fy:
            duplicates.append({"path": path, "reason": "no fiscal year found — skipped"})
            continue
        if fy not in chosen or len(os.path.basename(path)) < len(os.path.basename(chosen[fy])):
            if fy in chosen:
                duplicates.append({"path": chosen[fy], "reason": f"duplicate of {fy}"})
            chosen[fy] = path
        else:
            duplicates.append({"path": path, "reason": f"duplicate of {fy}"})

    results = []
    for fy in sorted(chosen):
        results.append(
            run(
                path=chosen[fy],
                company=company,
                commit=commit,
                unsynced_status=unsynced_status,
                with_print_log=with_print_log,
            )
        )

    created_key = "bills_" + ("created" if commit else "to_create")
    logs_key = "print_log_rows_" + ("created" if commit else "to_create")
    summary = {
        "dry_run": not commit,
        "years": sorted(chosen),
        "files_used": {fy: os.path.basename(p) for fy, p in sorted(chosen.items())},
        "files_skipped": duplicates,
        "sheet_rows_total": sum(r["sheet_rows"] for r in results),
        created_key: sum(r[created_key] for r in results),
        logs_key: sum(r[logs_key] for r in results),
        "skipped_already_have_a_bill": sum(r["skipped_already_have_a_bill"] for r in results),
        "skipped_invoice_not_in_erpnext": sum(
            r["skipped_invoice_not_in_erpnext"] for r in results
        ),
        "per_year": {
            r["fiscal_year"]: {
                "sheet_rows": r["sheet_rows"],
                created_key: r[created_key],
                logs_key: r[logs_key],
                "not_in_erpnext": r["skipped_invoice_not_in_erpnext"],
            }
            for r in results
        },
    }
    print("\n===== ALL YEARS =====")
    print(json.dumps(summary, indent=1, default=str))
    return summary
