# Copyright (c) 2026, Raindrop and contributors
# For license information, please see license.txt

"""One-off import of legacy print counts into Sales Invoice Print Count.

The old NGI billing software exports a "Sale Invoice Register" (.xls) listing,
per invoice, every print event with "No of Copy Printed". The migration stored
the old software's invoice number (e.g. NGI000001/82-83) in the Sales Invoice's
custom_branch_name.

That number identifies one invoice only within a COMPANY and a FISCAL YEAR — the
old ERPs numbered each company independently and restarted the counter every year
(NGK holds NGK/000001 in both 77/78 and 79/80). The NGI register format embeds
both (the NGI prefix and the /82-83 suffix), so its rows do resolve uniquely; a
register whose numbers carry neither does not. Pass company= to narrow the
lookup. Any register number that still matches several invoices is reported under
ambiguous_register_rows and imported for NONE of them, rather than crediting the
sheets to whichever invoice the scan happened to return last.

Per invoice this sums the copies over all its print events (the old software
counted the first print as 2 — the TAX INVOICE + INVOICE pair — same sheet
semantics as print_count.py) and upserts Sales Invoice Print Count:

  - no row yet  -> insert with print_count = legacy sheet total
  - row exists  -> ADD the legacy sheets to the current count (ERPNext prints
                   happened after the legacy ones; the counter is total sheets)

DRY RUN unless commit=True — always inspect the dry-run summary first.
NOT idempotent: running commit=True twice adds the legacy sheets twice.

The register exports ship in this folder; xls_path defaults to the 82-83 file.
The 80-81 and 81-82 files are proper full-year registers. The "79.80" file is
NOT a year register — it is an invoice-number-contains-"79" search export
(footer: "1478 of 100072") mixing old-format NGI/xxxxxx invoices with 580
80-81 invoices, and every invoice in it carries a phantom "ANIL, 1 copy" event,
all 1478 sharing the identical timestamp 2024-03-03 15:19:10 (one bulk action
fanned out, not real prints).

That batch event IS counted — a batch reprint produces real sheets, and for a
copy-number counter one sheet too many is safer than one too few (it can never
reuse a copy number). Pass drop_batch_events=True to exclude it instead; either
way the summary reports what was found.

Three knobs handle that file (and any future mixed export):

  - mode: "add" (default) adds this register's sheets to an existing counter;
    "max" raises the counter to this register's total instead. Use "max" when
    the register OVERLAPS one already imported — the 1478-row file shares 580
    invoices with the 80-81 register, and adding would count those twice.
    "max" is also idempotent, so re-running it is harmless;
  - only_prefix limits the import to invoice numbers starting with a prefix
    (e.g. "NGI/" for the old-format invoices only);
  - drop_batch_events excludes batch events from the totals.

DRY RUN unless commit=True — always inspect the dry-run summary first.
In mode="add", running commit=True twice adds the legacy sheets twice.

Usage (from the bench directory):

  # dry run — writes nothing, prints the summary:
  bench --site <site> execute \
    avinashgroup_app.legacy_print_import.import_legacy_print_counts.run

  # then, after checking the dry-run numbers:
  bench --site <site> execute \
    avinashgroup_app.legacy_print_import.import_legacy_print_counts.run \
    --kwargs "{'commit': True}"

  # the mixed "79.80" search export, AFTER the 80-81 register is in:
  bench --site <site> execute \
    avinashgroup_app.legacy_print_import.import_legacy_print_counts.run \
    --kwargs "{'xls_path': '<folder>/79.80 NGI.xls', 'mode': 'max'}"
"""

import json
import os
import re
from datetime import datetime, timedelta

import frappe
from frappe.utils import cint

from avinashgroup_app.utils.fiscal_year_utils import fiscal_year_for_date

REGISTER_FOLDER = os.path.join(os.path.dirname(__file__), "registers")
DEFAULT_XLS = os.path.join(REGISTER_FOLDER, "NGI", "82.83 NGI.xls")

# registers/<KEY>/ holds one company's exports.
#
# The "Sharwan Latest" exports in the same drop cover the live fiscal year
# 83/84, which ERPNext is also counting through print_count.py. That is only a
# double-count if the SAME invoice was printed in both systems, and under
# mode="max" not even then. All five were checked invoice-by-invoice against
# ng-group before being added as "83.84 Shrawan <X>.xls": they record prints
# the old software made while running alongside ERPNext through the
# 2026-07-17..2026-08-06 cutover, every number resolves, and NONE of them
# touch an invoice ERPNext had already printed (each company had only its
# first 83/84 invoice + return counted, GE 4 rows and the rest 2 each).
#
#   GE 567 inv/1064 sheets, NGI 481/853, NGG 299/586, NGK 283/572, NGN 1036/2127
#
# They are month views ending 2026-08-05/06, not full-year registers — re-pull
# and re-run to pick up later prints; mode="max" makes that safe.
#
# That re-pull happened on 2026-08-17, added as "83.84 <X> 2026-08-16.xls":
# the same five registers run forward to 2026-08-16, 4301 invoices/8415 sheets.
#
#   NGI 1000/1831, NGN 1476/3025, GE 953/1813, NGG 461/912, NGK 411/834
#
# The Shrawan files are KEPT rather than replaced. Four of the five new exports
# are strict supersets of their Shrawan counterpart, but GE's is not — it drops
# 28 invoices the Shrawan file holds (/SB000289-309, INV000243-250, 49 sheets),
# so both are needed and mode="max" merges them. NGK's new export is also the
# stripped 6-column layout with no "Printed by User" column, so only the
# Shrawan file can attribute those prints in run_print_log_backfill.
#
# 83 of the 4301 numbers do not resolve, all of them the tail of a series
# (NGI 28, GE 16, NGG 14, NGN 13, NGK 12) — invoices the old software printed
# on 2026-08-16 that ERPNext had not received yet when this was checked. 0
# ambiguous. They resolve on the next re-pull; mode="max" makes that safe too.
#
# Third pull 2026-08-18, NGI ONLY, "83.84 NGI 2026-08-18.xls": 1090 inv/2015
# sheets, 2026-07-17..2026-08-18. A strict superset of BOTH earlier NGI files
# (0 dropped, 0 lower), full 19-column layout with "Printed by User" intact.
# Delta over the live site: 11 inserts (21 sheets) + 2 raises (2 sheets), 0
# ambiguous. 107 unmatched — a contiguous tail from NGI004423 up, because the
# register runs to 08-18 while ERPNext's last 83/84 invoice is 2026-08-15.
# The other four companies were NOT re-pulled in this round.
REGISTER_COMPANIES = {
    "NGI": "Nepal Gas Udhyog Pvt. Ltd.",
    "NGG": "Nepal Gas Udhyog (Gandaki) Pvt. Ltd.",
    "NGK": "Nepal Gas Udhyog (Karnali) Pvt. Ltd.",
    "NGN": "Nepal Gas Udhyog (Narayani) Pvt. Ltd.",
    "GE": "Grishma Enterprises Pvt. Ltd.",
}

INVOICE_HEADER = "Invoice Number"
COPIES_HEADER = "No of Copy Printed"
DATE_HEADER = "Date & Time of Print"
USER_HEADER = "Printed by User"

# Excel serial-date epoch (datemode 0 workbooks, which these registers are).
EXCEL_EPOCH = datetime(1899, 12, 30)

# An exact print timestamp that repeats on more rows than this across the
# register is a batch event: one action logged against every invoice at the
# same instant (the "79.80" file has one such instant on 1478 invoices).
# Counted like any other print unless drop_batch_events is set — a batch
# reprint produces real sheets, and for a copy-number counter an extra sheet
# is the safe direction to err in.
BATCH_SHARE_LIMIT = 3


def excel_datetime(serial):
    """Excel serial date -> datetime, or None when the cell held no usable date."""
    if not isinstance(serial, (int, float)) or not serial:
        return None
    return EXCEL_EPOCH + timedelta(days=serial)


def _open_grid(path):
    """(nrows, ncols, cell(r, c)) for a register in either workbook format.

    Most registers are .xls and are read with xlrd. The NGI 79/80 register
    arrived as .xlsx, which xlrd 2.0 refuses outright ("Excel xlsx file; not
    supported"), so that one goes through openpyxl. Both are presented as the
    same read-only cell grid, and callers stay format-blind.

    openpyxl rows are ragged — a trailing empty cell is simply absent — so they
    are padded to the widest row; the register readers index fixed columns.
    """
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        ncols = max((len(r) for r in grid), default=0)
        grid = [r + [None] * (ncols - len(r)) for r in grid]
        return len(grid), ncols, lambda r, c: ("" if grid[r][c] is None else grid[r][c])

    import xlrd

    sheet = xlrd.open_workbook(path).sheet_by_index(0)
    return sheet.nrows, sheet.ncols, sheet.cell_value


def register_datetime(value):
    """Print timestamp from a register cell.

    The .xls registers store it as an Excel serial. The .xlsx one stores it as
    a real datetime (or a string), which excel_datetime cannot read — it would
    return None for every event and leave the whole file looking like one batch
    instant. Anything unrecognised is None, as before.
    """
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S",
                    "%m/%d/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None
    return excel_datetime(value)


def read_register_events(xls_path, drop_batch_events=False, info=None):
    """Every print event in a register export, in file order.

    Layout (both the 19- and 11-column exports): a header row names the
    columns; below it, an invoice row carries the invoice number in column A
    and each print event follows on its own row with the printing user, the
    copy count and the print timestamp in their named columns. Column
    POSITIONS differ between registers (the 81-82 file puts copies at 7 and the
    timestamp at 10, the others at 14 and 17), which is why they are located by
    header text rather than by index.

    Returns (events, invoices):
      events   - [{"invoice_no", "user", "timestamp", "copies"}], one per print
                 event, timestamp a datetime or None
      invoices - every invoice number the register lists, including those with
                 no print event at all

    Batch events (an exact timestamp shared by more than BATCH_SHARE_LIMIT
    rows) are reported in the info dict and, with drop_batch_events, left out
    of the returned events.
    """
    nrows, ncols, cell_value = _open_grid(xls_path)

    copies_col = date_col = user_col = header_row = None
    for r in range(nrows):
        for c in range(ncols):
            head = str(cell_value(r, c)).strip()
            if head == COPIES_HEADER:
                copies_col, header_row = c, r
            elif head == DATE_HEADER:
                date_col = c
            elif head == USER_HEADER:
                user_col = c
        if copies_col is not None:
            break
    if copies_col is None:
        frappe.throw(f"'{COPIES_HEADER}' header not found in {xls_path}")

    events = []
    invoices = []
    timestamp_rows = {}
    current = None
    for r in range(header_row + 1, nrows):
        invoice_no = str(cell_value(r, 0)).strip()
        if invoice_no == "Report Parameters":  # footer block ends the register
            break
        if invoice_no and invoice_no != INVOICE_HEADER:
            current = invoice_no
            invoices.append(current)  # register the invoice itself
            continue
        copies = cell_value(r, copies_col)
        if current and copies:
            serial = cell_value(r, date_col) if date_col is not None else None
            events.append(
                {
                    "invoice_no": current,
                    "user": (
                        str(cell_value(r, user_col)).strip()
                        if user_col is not None
                        else ""
                    ),
                    "timestamp": register_datetime(serial),
                    "copies": cint(copies),
                    "_serial": serial,
                }
            )
            if serial:
                timestamp_rows[serial] = timestamp_rows.get(serial, 0) + 1

    batch_ts = {ts for ts, n in timestamp_rows.items() if n > BATCH_SHARE_LIMIT}
    batch_events = sum(1 for e in events if e["_serial"] in batch_ts)
    batch_sheets = sum(e["copies"] for e in events if e["_serial"] in batch_ts)

    if drop_batch_events:
        events = [e for e in events if e["_serial"] not in batch_ts]

    if info is not None:
        info["batch_timestamps"] = [
            str(register_datetime(ts)) for ts in sorted(batch_ts, key=str)
        ]
        info["batch_events"] = batch_events
        info["batch_sheets"] = batch_sheets
        info["batch_events_dropped"] = drop_batch_events
    return events, invoices


def parse_register(xls_path, drop_batch_events=False, info=None):
    """Old invoice number -> total sheets printed, from a register export.

    The per-invoice view of read_register_events, which is what the print-count
    import needs. Invoices the register lists but never printed come back as 0.
    """
    events, invoices = read_register_events(
        xls_path, drop_batch_events=drop_batch_events, info=info
    )
    totals = {no: 0 for no in invoices}
    for e in events:
        totals.setdefault(e["invoice_no"], 0)
        totals[e["invoice_no"]] += e["copies"]
    return totals


# A register number that names its own fiscal year, e.g. NGI000001/82-83 or
# BHT000001/81-82. Anchored at the end so a number that merely contains digits
# cannot be mistaken for one. The separator varies — NGG's 79.80 register holds
# five numbers written NGG000001/80.81 with a dot.
NUMBER_FISCAL_YEAR = re.compile(r"(\d{2})[-/.](\d{2})\s*$")

# The fiscal year in a register FILENAME: "80.81 NGN.xls", "NGK 79.80.xls",
# "83.84 Shrawan NGI.xls".
FILE_FISCAL_YEAR = re.compile(r"(?<!\d)(\d{2})\.(\d{2})(?!\d)")


def register_fiscal_year(xls_path):
    """Fiscal year a register file covers, e.g. "79/80", or None.

    Taken from the FILENAME, not the sheet header: the header names the year
    the report was run under, which drifts — GE's 83/84 Shrawan export still
    says "FY [2082.083]" while every number in it is an 83-84 one.
    """
    match = FILE_FISCAL_YEAR.search(os.path.basename(xls_path))
    return "{0}/{1}".format(*match.groups()) if match else None


def number_fiscal_year(old_no, file_fiscal_year):
    """Fiscal year a single register number belongs to.

    A number that carries its own year wins — the year-register files are not
    pure, they spill into the next year (GE's 80.81 export holds 1367 bare
    80/81 numbers AND 376 explicit /81-82 ones, NGN's 79.80 export 9817 bare
    plus 6954 /80-81). A bare number belongs to the file's own year.
    """
    match = NUMBER_FISCAL_YEAR.search((old_no or "").strip())
    return "{0}/{1}".format(*match.groups()) if match else file_fiscal_year


def resolve_register_invoices(company=None):
    """(old invoice number, fiscal year) -> SI name.

    The number is only unique PER COMPANY PER FISCAL YEAR — the old ERPs
    numbered each company independently and restarted every year, so NGN/007208
    names four different invoices (76/77, 77/78, 78/79 and 79/80) and
    KTM/00001 three. Company comes from the company= argument; the year is what
    this key adds, and it is what makes the older bare-numbered registers
    resolvable at all.

    Returns (mapping, ambiguous), both keyed by that pair. A number still held
    by several invoices WITHIN one company and year lands in `ambiguous` and
    NOT in `mapping`, so callers skip it rather than resolving it by guessing —
    which a plain dict does silently by keeping whichever row the scan returned
    last. That guess is not hypothetical: it put 2,616 counters on NGN 79/80
    invoices on 2026-08-10, each holding the max of four unrelated invoices'
    print histories.
    """
    rows = frappe.db.sql(
        "SELECT custom_branch_name, name, company, posting_date "
        "FROM `tabSales Invoice` WHERE IFNULL(custom_branch_name, '') != ''"
        + (" AND company = %s" if company else ""),
        (company,) if company else (),
    )
    holders = {}
    for old_no, name, si_company, posting_date in rows:
        key = ((old_no or "").strip(), fiscal_year_for_date(posting_date))
        holders.setdefault(key, []).append(
            "{0} ({1})".format(name, si_company))
    mapping = {k: v[0].split(" (")[0] for k, v in holders.items() if len(v) == 1}
    ambiguous = {k: v for k, v in holders.items() if len(v) > 1}
    return mapping, ambiguous


def run(xls_path=None, commit=False, only_prefix=None, drop_batch_events=False, mode="add",
        company=None):
    if mode not in ("add", "max"):
        frappe.throw("mode must be 'add' or 'max'")
    xls_path = xls_path or DEFAULT_XLS
    info = {}
    totals = parse_register(xls_path, drop_batch_events=drop_batch_events, info=info)
    skipped_by_prefix = 0
    if only_prefix:
        skipped_by_prefix = sum(1 for no in totals if not no.startswith(only_prefix))
        totals = {no: n for no, n in totals.items() if no.startswith(only_prefix)}

    mapping, ambiguous = resolve_register_invoices(company)
    file_fiscal_year = register_fiscal_year(xls_path)

    existing = {
        r.name: cint(r.print_count)
        for r in frappe.get_all(
            "Sales Invoice Print Count", fields=["name", "print_count"]
        )
    }

    inserts, updates, unmatched, ambiguous_hits = [], [], [], []
    for old_no, sheets in totals.items():
        if not sheets:
            continue
        key = (old_no, number_fiscal_year(old_no, file_fiscal_year))
        si = mapping.get(key)
        if not si and key in ambiguous:
            # several invoices hold this number in this very year: counting it
            # would credit sheets to an arbitrary one. Report it and import
            # nothing for this row.
            ambiguous_hits.append({"old_no": old_no, "fiscal_year": key[1],
                                   "sheets": sheets,
                                   "candidates": ambiguous[key][:5]})
        elif not si:
            unmatched.append(old_no)
        elif si in existing:
            was = existing[si]
            to = max(was, sheets) if mode == "max" else was + sheets
            if to == was:  # 'max' and the counter already covers this register
                continue
            updates.append(
                {
                    "invoice": si,
                    "old_no": old_no,
                    "from": was,
                    "add": to - was,
                    "to": to,
                }
            )
        else:
            inserts.append({"invoice": si, "old_no": old_no, "print_count": sheets})

    if commit:
        for row in inserts:
            frappe.get_doc(
                {
                    "doctype": "Sales Invoice Print Count",
                    "sales_invoice": row["invoice"],
                    "branch_name": row["old_no"],
                    "print_count": row["print_count"],
                }
            ).insert(ignore_permissions=True)
        for row in updates:
            frappe.db.sql(
                "UPDATE `tabSales Invoice Print Count` "
                "SET print_count = print_count + %s WHERE name = %s",
                (row["add"], row["invoice"]),
            )
        frappe.db.commit()

    result = {
        "dry_run": not commit,
        "xls_path": xls_path,
        "mode": mode,
        "only_prefix": only_prefix,
        "skipped_by_prefix": skipped_by_prefix,
        "company": company or "(all companies)",
        **info,
        "excel_invoices": len(totals),
        ("inserted" if commit else "would_insert"): len(inserts),
        ("updated" if commit else "would_update"): len(updates),
        "unmatched": len(unmatched),
        "unmatched_invoices": unmatched[:50],
        "fiscal_year": file_fiscal_year,
        "ambiguous_numbers_in_db": len(ambiguous),
        "ambiguous_register_rows": len(ambiguous_hits),
        "ambiguous_detail": ambiguous_hits[:20],
        "total_sheets": sum(r["print_count"] for r in inserts)
        + sum(r["add"] for r in updates),
    }
    if not file_fiscal_year:
        result["fiscal_year_warning"] = (
            "No fiscal year in the filename, so bare register numbers (those not "
            "carrying their own /82-83 suffix) cannot be placed in a year and will "
            "look ambiguous. Rename the file to include it, e.g. '80.81 NGN.xls'."
        )
    if ambiguous_hits:
        result["ambiguous_warning"] = (
            "{0} register row(s) name a number held by SEVERAL invoices of the SAME "
            "company AND fiscal year, and were SKIPPED. Company and year already "
            "narrow the lookup, so this is a genuine duplicate in the data rather "
            "than the year-restart collision — inspect ambiguous_detail.".format(
                len(ambiguous_hits))
        )
    if updates:
        result["update_rows"] = updates[:50]
        result["warning"] = (
            "updates ADD legacy sheets to existing counters - if this import "
            "already ran once, committing again will double-count"
            if mode == "add"
            else "updates RAISE counters to this register's total; re-running "
            "changes nothing"
        )
    print(json.dumps(result, indent=1, default=str))
    return result


def backfill_fiscal_year(commit=False, company=None):
    """Fill Sales Invoice Print Count.fiscal_year on rows written before it existed.

    The field mirrors the invoice's posting date, so this is pure derivation —
    it changes no count. Rows whose invoice has no Fiscal Year row spanning its
    posting date are reported and left alone.

    DRY RUN unless commit=True.

      bench --site <site> execute \\
        avinashgroup_app.legacy_print_import.import_legacy_print_counts.backfill_fiscal_year \\
        --kwargs "{'commit': True}"
    """
    filters = {"fiscal_year": ["in", ["", None]]}
    if company:
        filters["company"] = company
    rows = frappe.get_all(
        "Sales Invoice Print Count", filters=filters, fields=["name", "sales_invoice"]
    )
    dates = dict(
        frappe.db.sql(
            "SELECT name, posting_date FROM `tabSales Invoice` "
            "WHERE IFNULL(custom_branch_name, '') != ''"
            + (" AND company = %s" if company else ""),
            (company,) if company else (),
        )
    )
    by_year, no_year = {}, []
    for row in rows:
        year = fiscal_year_for_date(dates.get(row.sales_invoice))
        if year:
            by_year.setdefault(year, []).append(row.name)
        else:
            no_year.append(row.name)

    if commit:
        for year, names in by_year.items():
            for chunk in _chunks(names, 500):
                frappe.db.sql(
                    "UPDATE `tabSales Invoice Print Count` SET fiscal_year = %s "
                    "WHERE name IN ({0})".format(", ".join(["%s"] * len(chunk))),
                    [year] + list(chunk),
                )
        frappe.db.commit()

    result = {
        "dry_run": not commit,
        "company": company or "(all companies)",
        "counters_without_fiscal_year": len(rows),
        ("updated" if commit else "would_update"): sum(
            len(v) for v in by_year.values()
        ),
        "by_fiscal_year": {k: len(v) for k, v in sorted(by_year.items())},
        "no_fiscal_year_row_for_posting_date": len(no_year),
        "no_fiscal_year_sample": no_year[:20],
    }
    print(json.dumps(result, indent=1, default=str))
    return result


PRINT_LOG_FIELDS = (
    "name",
    "creation",
    "modified",
    "modified_by",
    "owner",
    "docstatus",
    "idx",
    "sales_invoice",
    "customer",
    "customer_name",
    "branch_name",
    "company",
    "copy_number",
    "printed_by",
)


def run_print_log_backfill(xls_path=None, commit=False, only_prefix=None,
                           drop_batch_events=False, company=None, limit=None):
    """Recreate the old software's print history as Sales Invoice Print Log rows.

    run() above imports only the per-invoice sheet TOTAL, into Sales Invoice
    Print Count. The register also names, for every print event, who printed it
    and when — the "Printed by User" and "Date & Time of Print" columns — and
    that is what the Materialized Report's Printed / Printed Time / Printed By
    columns and the Invoice Activity Report's audit trail read. This replays
    those events into the Print Log.

    One row per SHEET, matching print_count.py: an event that printed 2 copies
    becomes two rows sharing its timestamp and user, and copy_number runs
    1..n across all of an invoice's events in chronological order.

    The register's user names (ASHISH, NDILLI) go into `printed_by` verbatim.
    They are names in the OLD software, not Frappe users — ASHISH alone matches
    two live accounts — so nothing is mapped to a User and `owner` stays the
    account running the import.

    IDEMPOTENT: an invoice that already has any Print Log row is skipped whole,
    so a second run inserts nothing and a partially-logged invoice is never
    given a duplicate sheet history.

    DRY RUN unless commit=True — always inspect the dry-run summary first.

    Usage (from the bench directory):

      bench --site <site> execute \\
        avinashgroup_app.legacy_print_import.import_legacy_print_counts.run_print_log_backfill

      bench --site <site> execute \\
        avinashgroup_app.legacy_print_import.import_legacy_print_counts.run_print_log_backfill \\
        --kwargs "{'commit': True}"
    """
    xls_path = xls_path or DEFAULT_XLS
    info = {}
    events, _invoices = read_register_events(
        xls_path, drop_batch_events=drop_batch_events, info=info
    )

    skipped_by_prefix = 0
    if only_prefix:
        before = len(events)
        events = [e for e in events if e["invoice_no"].startswith(only_prefix)]
        skipped_by_prefix = before - len(events)

    # An event with no usable timestamp cannot be placed in the sheet order and
    # would have to invent a `creation` — report it instead of guessing.
    undated = [e for e in events if not e["timestamp"]]
    events = [e for e in events if e["timestamp"]]

    mapping, ambiguous = resolve_register_invoices(company)
    file_fiscal_year = register_fiscal_year(xls_path)

    by_invoice = {}
    unmatched, ambiguous_hits = set(), []
    seen_ambiguous = set()
    for e in events:
        old_no = e["invoice_no"]
        key = (old_no, number_fiscal_year(old_no, file_fiscal_year))
        si = mapping.get(key)
        if not si:
            if key in ambiguous:
                if key not in seen_ambiguous:
                    seen_ambiguous.add(key)
                    ambiguous_hits.append({"old_no": old_no, "fiscal_year": key[1],
                                           "candidates": ambiguous[key][:5]})
            else:
                unmatched.add(old_no)
            continue
        by_invoice.setdefault(si, []).append(e)

    already_logged = _invoices_with_print_log(list(by_invoice))
    skipped_existing = [si for si in by_invoice if si in already_logged]
    for si in skipped_existing:
        del by_invoice[si]

    targets = sorted(by_invoice)
    if limit:
        targets = targets[: int(limit)]

    details = _invoice_details(targets)
    now = frappe.utils.now()
    user = frappe.session.user

    rows, sample = [], []
    for si in targets:
        d = details.get(si) or {}
        sheet = 0
        for e in sorted(by_invoice[si], key=lambda e: e["timestamp"]):
            for _copy in range(e["copies"]):
                sheet += 1
                rows.append(
                    (
                        frappe.generate_hash(length=10),
                        e["timestamp"],
                        now,
                        user,
                        user,
                        0,
                        0,
                        si,
                        d.get("customer"),
                        d.get("customer_name"),
                        d.get("custom_branch_name"),
                        d.get("company"),
                        sheet,
                        e["user"] or None,
                    )
                )
                if len(sample) < 5:
                    sample.append(
                        {
                            "invoice": si,
                            "branch_name": d.get("custom_branch_name"),
                            "copy_number": sheet,
                            "creation": str(e["timestamp"]),
                            "printed_by": e["user"],
                        }
                    )

    if commit and rows:
        frappe.db.bulk_insert("Sales Invoice Print Log", list(PRINT_LOG_FIELDS), rows)
        frappe.db.commit()

    result = {
        "dry_run": not commit,
        "xls_path": xls_path,
        "only_prefix": only_prefix,
        "skipped_by_prefix": skipped_by_prefix,
        "company": company or "(all companies)",
        **info,
        "register_events": len(events),
        "undated_events_skipped": len(undated),
        "invoices_matched": len(by_invoice) + len(skipped_existing),
        "invoices_already_logged_skipped": len(skipped_existing),
        "invoices_to_backfill": len(targets),
        ("inserted_rows" if commit else "would_insert_rows"): len(rows),
        "unmatched_numbers": len(unmatched),
        "unmatched_sample": sorted(unmatched)[:50],
        "ambiguous_register_rows": len(ambiguous_hits),
        "ambiguous_detail": ambiguous_hits[:20],
        "printers": _printer_tally(by_invoice, targets),
        "sample_rows": sample,
    }
    if ambiguous_hits:
        result["ambiguous_warning"] = (
            "{0} register row(s) name a number held by SEVERAL invoices and were "
            "SKIPPED — the number is only unique per company per fiscal year. "
            "Re-run with company='<the register's company>'.".format(len(ambiguous_hits))
        )
    if limit and len(by_invoice) > len(targets):
        result["limit_warning"] = (
            "limit={0} applied: {1} matched invoices were NOT backfilled this "
            "run".format(limit, len(by_invoice) - len(targets))
        )
    print(json.dumps(result, indent=1, default=str))
    return result


def _invoices_with_print_log(invoices):
    """Subset of `invoices` that already has at least one Print Log row."""
    if not invoices:
        return set()
    found = set()
    for chunk in _chunks(invoices, 5000):
        found.update(
            r[0]
            for r in frappe.db.sql(
                "SELECT DISTINCT sales_invoice FROM `tabSales Invoice Print Log` "
                "WHERE sales_invoice IN %s",
                (tuple(chunk),),
            )
        )
    return found


def _invoice_details(invoices):
    """customer / customer_name / branch_name / company per invoice, for the Log
    row's denormalized columns (print_count.py stores the same snapshot)."""
    details = {}
    for chunk in _chunks(invoices, 5000):
        for r in frappe.db.sql(
            "SELECT name, customer, customer_name, custom_branch_name, company "
            "FROM `tabSales Invoice` WHERE name IN %s",
            (tuple(chunk),),
            as_dict=True,
        ):
            details[r.name] = r
    return details


def _printer_tally(by_invoice, targets):
    """How many sheets each register user accounts for in this run — the quickest
    check that printed_by came through rather than landing empty."""
    tally = {}
    for si in targets:
        for e in by_invoice[si]:
            name = e["user"] or "(blank)"
            tally[name] = tally.get(name, 0) + e["copies"]
    return dict(sorted(tally.items(), key=lambda kv: kv[1], reverse=True))


def _chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i : i + size]


def run_group(folder=REGISTER_FOLDER, companies=None, commit=False,
              drop_batch_events=False, with_print_log=False):
    """Import every register in registers/<KEY>/ — the whole group, one command.

    Each company's registers are resolved against that company's invoices, which
    is what makes the legacy numbers unambiguous: NGK/000001 exists in several
    fiscal years and NGG/KTM numbers repeat across companies.

    ALWAYS mode="max". A register lists every print event of every invoice it
    names, so for any one invoice a single register already holds the complete
    total; "add" would double-count the invoices two registers share. Several
    files here are search exports rather than year registers and overlap heavily
    (NGN 79/80 carries 16,771 invoices against that year's 9,823), and NGI's
    registers were already imported once on ng-group. "max" is idempotent, so
    this command is safe to repeat and safe to resume.

    DRY RUN unless commit=True.

      bench --site <site> execute \\
        avinashgroup_app.legacy_print_import.import_legacy_print_counts.run_group \\
        --kwargs "{'commit': True}"
    """
    import glob

    keys = companies or sorted(REGISTER_COMPANIES)
    results, missing = [], []
    for key in keys:
        company = REGISTER_COMPANIES.get(key)
        if not company:
            frappe.throw(f"Unknown register folder '{key}'")
        paths = sorted(glob.glob(os.path.join(folder, key, "*.xls*")))
        if not paths:
            missing.append(key)
            continue
        for path in paths:
            out = run(
                xls_path=path,
                commit=commit,
                drop_batch_events=drop_batch_events,
                mode="max",
                company=company,
            )
            out["_file"] = f"{key}/{os.path.basename(path)}"
            out["_company"] = company
            results.append(out)
            if with_print_log:
                log = run_print_log_backfill(
                    xls_path=path, commit=commit, company=company
                )
                log["_file"] = out["_file"]
                results.append(log)

    def total(*fields):
        return sum(r.get(f) or 0 for r in results for f in fields)

    inserted_key = "counters_inserted" if commit else "counters_to_insert"
    updated_key = "counters_raised" if commit else "counters_to_raise"
    summary = {
        "dry_run": not commit,
        "mode": "max",
        "folders_empty": missing,
        "files": [r["_file"] for r in results],
        "register_invoices": total("excel_invoices"),
        inserted_key: total("inserted", "would_insert"),
        updated_key: total("updated", "would_update"),
        "sheets_applied": total("total_sheets"),
        "unmatched_register_rows": total("unmatched"),
        "ambiguous_register_rows": total("ambiguous_register_rows"),
        "batch_events_counted": total("batch_events"),
        "per_file": [
            {
                "file": r["_file"],
                "register_invoices": r.get("excel_invoices"),
                "insert": r.get("inserted", r.get("would_insert")),
                "raise": r.get("updated", r.get("would_update")),
                "sheets": r.get("total_sheets"),
                "unmatched": r.get("unmatched"),
                "ambiguous": r.get("ambiguous_register_rows"),
            }
            for r in results
        ],
    }
    print(json.dumps(summary, indent=1, default=str))
    return summary
