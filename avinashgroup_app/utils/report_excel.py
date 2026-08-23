# Copyright (c) 2026, Raindrop and contributors
# For license information, please see license.txt

"""Excel downloads with a company letterhead on top of the table.

Three entry points:
- send_report_xlsx: used by Script Reports (Materialized Return) to build the
  whole workbook, including a totals row for numeric columns.
- send_annexure7_xlsx: the Materialized Report's own layout, reproducing the
  VAT Annexure 7 sheet the old NGI billing software exported.
"""

from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt
from frappe.utils.xlsxutils import make_xlsx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def get_company_header_rows(company):
	"""Letterhead rows pulled from the Company record and its linked Addresses.

	    [company name]                                   <- bolded by make_xlsx (row 1)
	    [Pan No.: <tax_id>]
	    [<address line1, city>, Ph: <phone>, Factory: <factory line1, city>]

	Every line — and every fragment of the address line — is dropped when the
	company has no data for it, so nothing renders as an empty label.
	"""
	if not company:
		return []

	company_doc = frappe.db.get_value(
		"Company", company, ["company_name", "tax_id", "phone_no"], as_dict=True
	)
	if not company_doc:
		return []

	rows = [[company_doc.company_name or company]]

	if company_doc.tax_id:
		rows.append(["Pan No.: {0}".format(company_doc.tax_id)])

	addresses = frappe.db.sql(
		"""
		select a.address_line1, a.city, a.phone, a.address_type, a.address_title
		from `tabAddress` a
		join `tabDynamic Link` dl on dl.parent = a.name and dl.parenttype = 'Address'
		where dl.link_doctype = 'Company' and dl.link_name = %s and a.disabled = 0
		order by a.is_primary_address desc, a.creation asc
		""",
		company,
		as_dict=True,
	)

	def address_text(addr):
		return ", ".join(filter(None, [addr.address_line1, addr.city]))

	def is_factory(addr):
		return "factory" in ((addr.address_type or "") + (addr.address_title or "")).lower()

	main = next((a for a in addresses if not is_factory(a)), None)
	factory = next((a for a in addresses if is_factory(a)), None)

	parts = []
	if main and address_text(main):
		parts.append(address_text(main))

	phone = company_doc.phone_no or (main and main.phone)
	if phone:
		parts.append("Ph: {0}".format(phone))

	if factory and address_text(factory):
		parts.append("Factory: {0}".format(address_text(factory)))

	if parts:
		rows.append([", ".join(parts)])

	return rows


def send_report_xlsx(columns, data, company, title, filename):
	"""Build the workbook and attach it to frappe.response as a file download.

	columns/data are exactly what the report's execute() returned; Check columns
	come out as Excel TRUE/FALSE and Float columns get a summed totals row.
	"""
	rows = get_company_header_rows(company)
	rows.append([title])
	letterhead_row_count = len(rows)
	rows.append([c["label"] for c in columns])

	# execute() pads an empty result with one blank dict so the desk keeps its
	# column headers visible — drop that padding here.
	data = [d for d in data if isinstance(d, dict) and d]

	float_fields = [c["fieldname"] for c in columns if c["fieldtype"] == "Float"]
	totals = {f: 0 for f in float_fields}
	for d in data:
		row = []
		for c in columns:
			value = d.get(c["fieldname"])
			if c["fieldtype"] == "Check":
				value = bool(value)
			row.append(value)
		rows.append(row)
		for f in float_fields:
			totals[f] += flt(d.get(f))

	if data:
		total_row = [
			flt(totals[c["fieldname"]], 2) if c["fieldname"] in totals else "" for c in columns
		]
		total_row[0] = _("Total")
		rows.append(total_row)

	xlsx = make_xlsx(rows, title.title())
	xlsx = _center_letterhead(xlsx, letterhead_row_count, len(columns))
	frappe.response.filename = filename
	frappe.response.filecontent = xlsx.getvalue()
	frappe.response.type = "binary"


ANNEXURE7_TOTAL_FIELDS = ("taxable_amount", "tax_amount", "total_amount")

# Indian digit grouping (#,##,##0.00 -> 2,95,82,90,259.83), as the legacy export uses.
ANNEXURE7_NUMBER_FORMAT = "#,##,##0.00"

# Column widths in Excel character units, copied from the legacy export.
ANNEXURE7_COLUMN_WIDTHS = {
	1: 17.14, 2: 25.71, 3: 22.85, 4: 28.57, 5: 17.14,
	6: 21.42, 7: 21.42, 8: 21.42, 9: 21.42, 10: 21.42, 11: 21.42,
	12: 15.71, 13: 15.71, 14: 28.57,
	15: 21.42, 16: 21.42, 17: 21.42, 18: 21.42,
	19: 28.57, 20: 28.57, 21: 28.57,
}

# Left-aligned data columns; everything else that is not a number is centered.
ANNEXURE7_LEFT_ALIGNED = {"customer_name", "transaction_id"}


def send_annexure7_xlsx(columns, data, company, filename, fy_display="", period="", date_range=""):
	"""Build the VAT Annexure 7 workbook and attach it as a file download.

	Reproduces the layout of the old NGI billing software's export: a five-row
	letterhead block (company + fiscal year left, annexure title and accounting
	period right), the column headers on row 6, and a totals row under the three
	amount columns the annexure totals.

	Built in one pass rather than written-then-reopened like send_report_xlsx:
	a full fiscal year is ~37,000 rows, and loading that back to add merges
	would double the memory for nothing.
	"""
	from openpyxl import Workbook

	wb = Workbook()
	ws = wb.active
	ws.title = "Annexure-7"
	width = len(columns)

	_annexure7_letterhead(ws, company, width, fy_display, period, date_range)
	header_row = 6

	for idx, col in enumerate(columns, 1):
		cell = ws.cell(row=header_row, column=idx, value=col["label"])
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
		ws.column_dimensions[cell.column_letter].width = ANNEXURE7_COLUMN_WIDTHS.get(idx, 21.42)
	ws.row_dimensions[header_row].height = 20

	# execute() pads an empty result with one blank dict so the desk keeps its
	# column headers visible — drop that padding here.
	data = [d for d in data if isinstance(d, dict) and d]

	totals = {f: 0 for f in ANNEXURE7_TOTAL_FIELDS}
	number_align = Alignment(horizontal="right", vertical="bottom")
	text_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
	left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

	row_idx = header_row
	for d in data:
		row_idx += 1
		for idx, col in enumerate(columns, 1):
			fieldname = col["fieldname"]
			value = d.get(fieldname)
			if col["fieldtype"] == "Float":
				cell = ws.cell(row=row_idx, column=idx, value=flt(value))
				cell.number_format = ANNEXURE7_NUMBER_FORMAT
				cell.alignment = number_align
				if fieldname in totals:
					totals[fieldname] += flt(value)
			elif value is None or value == "":
				# Write no cell at all rather than a blank string — the columns
				# with no source (Payment Method, VAT Refund, Transaction Id)
				# are absent from the legacy export too, not empty-valued.
				continue
			else:
				cell = ws.cell(row=row_idx, column=idx, value=value)
				cell.alignment = left_align if fieldname in ANNEXURE7_LEFT_ALIGNED else text_align

	if data:
		row_idx += 1
		for idx, col in enumerate(columns, 1):
			fieldname = col["fieldname"]
			if idx == 5:
				cell = ws.cell(row=row_idx, column=idx, value=_("Totals"))
				cell.font = Font(bold=True)
				cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
			elif fieldname in totals:
				cell = ws.cell(row=row_idx, column=idx, value=flt(totals[fieldname], 2))
				cell.font = Font(bold=True)
				cell.number_format = ANNEXURE7_NUMBER_FORMAT
				cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

	out = BytesIO()
	wb.save(out)
	frappe.response.filename = filename
	frappe.response.filecontent = out.getvalue()
	frappe.response.type = "binary"


def _annexure7_letterhead(ws, company, width, fy_display, period, date_range):
	"""Rows 1-5: company block merged across A:E, annexure block across F:H,
	then a full-width report-parameters line on row 5."""
	rows = get_company_header_rows(company)
	company_name = rows[0][0] if rows else ""
	left_lines = [
		"{0} FY [{1}]".format(company_name, fy_display) if fy_display else company_name,
		rows[1][0] if len(rows) > 1 else "",
		rows[2][0] if len(rows) > 2 else "",
	]
	right_lines = [
		# The legacy export printed "VAT Annexure 7" here. Left blank on purpose:
		# the sheet is not being presented as that document. The row stays so the
		# accounting period and the currency note keep their positions — and so
		# the block still balances the company name opposite it.
		"",
		"{0} : {1}".format(_("Accounting Period"), period) if period else "",
		_("Amount in Nepalese Rupee (NPR)"),
	]

	title_font = Font(bold=True, size=14, name="Segoe UI Light", color="FF0000FF")
	right_title_font = Font(bold=True, size=14, name="Segoe UI Light")
	small_font = Font(size=10)

	for i in range(3):
		row = i + 1
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
		left = ws.cell(row=row, column=1, value=left_lines[i])
		left.font = title_font if i == 0 else small_font
		left.alignment = Alignment(horizontal="left", vertical="center")

		ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
		right = ws.cell(row=row, column=6, value=right_lines[i])
		right.font = right_title_font if i == 0 else small_font
		right.alignment = Alignment(horizontal="right", vertical="center")

	ws.row_dimensions[1].height = 30

	ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=width)
	params = ws.cell(
		row=5,
		column=1,
		value="{0} : {1} : {2}".format(_("Report Parameters"), _("Date Range"), date_range)
		if date_range
		else "",
	)
	params.font = small_font
	params.alignment = Alignment(horizontal="left", vertical="center")


def _center_letterhead(xlsx_file, row_count, table_width):
	"""Merge each letterhead row across the table's columns and center it.
	make_xlsx works on a write-only workbook, which cannot merge cells — so the
	styling is applied by reopening the finished file."""
	wb = load_workbook(xlsx_file)
	ws = wb.active
	for row in range(1, row_count + 1):
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=table_width)
		cell = ws.cell(row=row, column=1)
		cell.alignment = Alignment(horizontal="center")
		# Company name (row 1) and the report title (last letterhead row) in bold.
		if row == 1 or row == row_count:
			cell.font = Font(bold=True)
	out = BytesIO()
	wb.save(out)
	return out
